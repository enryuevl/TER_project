import os
from datetime import date
import re
from typing import Dict, Any, Optional

try:
    import customtkinter as ctk
except Exception as _e:
    # Allow importing this module in non-GUI contexts (e.g., unit tests)
    ctk = None

try:
    from openpyxl import load_workbook
    from openpyxl.utils.protection import hash_password
    from openpyxl.workbook.protection import WorkbookProtection
except Exception:
    load_workbook = None
    hash_password = None
    WorkbookProtection = None

try:
    from tkinter import messagebox
except Exception:
    messagebox = None



def _safe_filename(name: str) -> str:
    name = re.sub(r'[\\\\/:*?"<>|]+', "", name)
    name = re.sub(r"\\s+", " ", name).strip()
    return name


# ---------- Workbook protection helpers ----------
def _apply_workbook_protection(wb, password: str = "ATSLock"):
    """Protect workbook structure and all sheets with a password to deter tampering."""
    if not wb or not hash_password or not WorkbookProtection:
        return

    pw = password or "ATSLock"
    hashed = hash_password(pw)
    wb.security = WorkbookProtection(lockStructure=True, workbookPassword=hashed)

    for ws in wb.worksheets:
        try:
            ws.protection.sheet = True
            # set_password hashes internally; fallback to hashed if set_password missing
            if hasattr(ws.protection, "set_password"):
                ws.protection.set_password(pw)
            else:
                ws.protection.password = hashed
            ws.protection.enable()
        except Exception:
            continue


def protect_workbook_file(path: str, password: str = "ATSLock") -> str:
    """Load, protect, and save an existing workbook file. Returns the path."""
    if load_workbook is None:
        return path
    if not path or not os.path.exists(path):
        return path
    try:
        wb = load_workbook(path)
        _apply_workbook_protection(wb, password)
        wb.save(path)
    except Exception:
        pass
    return path


# ---------------- Core calculations (shared) ---------------- #

def compute_equivalent_numerical(overall_score: float) -> int:
    if overall_score >= 9.3:
        return 10
    elif overall_score >= 7.5:
        return 8
    elif overall_score >= 5:
        return 6
    elif overall_score >= 3:
        return 4
    else:
        return 2


def compute_adjective_rating(eq_num: int) -> str:
    mapping = {
        10: "Outstanding (O)",
        8:  "Very Satisfactory (VS)",
        6:  "Satisfactory (S)",
        4:  "Fair (F)",
        2:  "Unsatisfactory (US)"
    }
    return mapping.get(eq_num, "")


def calculate_row_averages(processed_results: Dict[str, Any], teacher_name: str, rater_type: Optional[str] = None) -> Dict[str, float]:
    teacher_data = processed_results.get(teacher_name, {})

    if isinstance(teacher_data, list):
        teacher_data = {"Unknown": teacher_data}

    if not rater_type:
        # Pick first rater if not provided
        rater_type = next(iter(teacher_data.keys()), None)

    if not rater_type or rater_type not in teacher_data:
        return {}

    row_scores: Dict[str, list] = {}

    # Collect rows across documents
    for _, result, *_ in teacher_data[rater_type]:
        for section, rows in result.items():
            for rownum, score in rows.items():
                row_key = f"{section} R{rownum}"
                row_scores.setdefault(row_key, []).append(score)

    # Average per row
    return {rk: (sum(vals) / len(vals)) for rk, vals in row_scores.items() if vals}


def infer_ay_and_sem_from_today() -> tuple[str, str]:
    """Infer AY and Sem from today's date (PH academic calendar)."""
    today = date.today()
    y, m = today.year, today.month
    if 8 <= m <= 12:
        return f"{y}-{y+1}", "1st Sem"
    elif 1 <= m <= 5:
        return f"{y-1}-{y}", "2nd Sem"
    else:  # Jun–Jul
        return f"{y-1}-{y}", "Midyear"


def ay_for_sem(selected_sem: str, anchor: Optional[date] = None) -> str:
    d = anchor or date.today()
    y = d.year
    if selected_sem == "1st Sem":
        return f"{y}-{y+1}"
    else:
        return f"{y-1}-{y}"


# ---------------- UI controller ---------------- #

class SummaryFormController:
    
    def __init__(self, processed_results: Dict[str, Any], db_module=None):
        self.processed_results = processed_results
        self.db = db_module
        self.entry_widgets: Dict[str, Any] = {}
        self.current_teacher: Optional[str] = None
        self.semester_var = None
        self.academic_year_var = None
        self.supervisor_override: Optional[dict] = None

    # ---------- DB lookup ----------
    def get_department_for_teacher(self, teacher_name: str) -> str:
        if not self.db:
            return ""
        try:
            conn = self.db.connect()
            cur = conn.cursor()
            cur.execute(
                """
                SELECT d.name
                FROM faculty f
                JOIN departments d ON f.department_id = d.id
                WHERE f.full_name = ?
                """,
                (teacher_name,),
            )
            row = cur.fetchone()
            conn.close()
            return row[0] if row else ""
        except Exception:
            return ""

    # ---------- Public API ----------
    def show(self, master, teacher_name: str):
        """Open the Summary popup and prefill values."""
        if ctk is None:
            raise RuntimeError("customtkinter is required to show the Summary UI.")

        self.current_teacher = teacher_name
        self._open_summary_popup(master, teacher_name)

        # Instructor + Department
        if "instructor" in self.entry_widgets:
            self.entry_widgets["instructor"].delete(0, "end")
            self.entry_widgets["instructor"].insert(0, teacher_name)

        dept = self.get_department_for_teacher(teacher_name)
        if dept and "department" in self.entry_widgets:
            self.entry_widgets["department"].delete(0, "end")
            self.entry_widgets["department"].insert(0, dept)

        # rater types → entry key → weight %
        mapping = {
            "Student": ("student_rater", 25),
            "Peer":    ("peer_rater",    15),
            "Dean":    ("dean_rater",    15),
        }
        for rater_key, (entry_key, weight) in mapping.items():
            averages = calculate_row_averages(self.processed_results, teacher_name, rater_key)
            grand_total = sum(averages.values()) if averages else 0.0

            self._set_value(entry_key, f"{grand_total:.2f}", field="rating")
            rating_equiv = grand_total / 10 if grand_total else 0.0
            self._set_value(entry_key, f"{rating_equiv:.2f}", field="equivalent")
            point_score = rating_equiv * (weight / 100.0)
            self._set_value(entry_key, f"{point_score:.2f}", field="point")

        # compute totals
        self._update_overall_point_score()

    # ---------- Export (optional) ----------
    def export_full_summary(self, template_path="template.xlsx") -> Optional[str]:
        if load_workbook is None:
            raise RuntimeError("openpyxl is required for export_full_summary.")

        # Resolve AY & Sem
        sem = None
        ay = None
        if self.semester_var is not None:
            try:
                sem = self.semester_var.get()
            except Exception:
                pass
        if self.academic_year_var is not None:
            try:
                ay = self.academic_year_var.get()
            except Exception:
                pass

        if sem and not ay:
            ay = ay_for_sem(sem)
        if not sem and not ay:
            ay, sem = infer_ay_and_sem_from_today()

        teacher_name = self.current_teacher or "Unknown Teacher"

        # Ensure Rating Period field reflects the computed values
        if "rating period" in self.entry_widgets:
            e = self.entry_widgets["rating period"]
            try:
                if self._widget_alive(e):
                    e.delete(0, "end")
                    e.insert(0, f"{sem} AY {ay}")
            except Exception:
                pass

        # Load template sheets
        wb = load_workbook(template_path)
        ws_ti = wb["TI"]
        ws_ter = wb["TER"]
        try:
            ws_entry = wb["Entry"]
        except Exception:
            ws_entry = None

        # Write data into sheets
        self._write_student_scores(ws_ti)
        self._write_manual_results(ws_ter)
        self._write_peer_scores(ws_ti)
        self._write_self_scores(ws_ti)
        self._write_dean_scores(ws_ti)

        # Populate header metadata for first-time/overwrite writes
        meta = self._get_teacher_meta(teacher_name)
        override = getattr(self, "supervisor_override", None) or None
        if override:
            meta["dean_name"] = override.get("name", meta.get("dean_name", ""))
            meta["dean_title"] = override.get("title", meta.get("dean_title", "Dean"))
        try:
            period_text = f"{sem} AY {ay}"
            if ws_entry is not None:
                ws_entry["D7"].value = period_text
                ws_entry["D12"].value = meta["first_name"]
                ws_entry["D13"].value = meta["last_name"]
                ws_entry["D14"].value = meta["middle_name"]
                ws_entry["D16"].value = meta["rank"]
                ws_entry["D18"].value = meta["subrank"]
                ws_entry["D20"].value = meta["department"]
                ws_entry["D23"].value = meta["dean_name"]
                ws_entry["D24"].value = meta.get("dean_title", "Dean")
        except Exception:
            pass

        # If the teacher being evaluated is a dean, align behavior weights on TI/TER (J25-J28) to 5%
        try:
            is_dean = self._is_teacher_dean(teacher_name)
        except Exception:
            is_dean = False
        if is_dean:
            for r in range(25, 29):  # rows 25-28, column J (10)
                cell = ws_ti.cell(row=r, column=10)
                cell.value = 0.05
                try:
                    cell.number_format = "0%"
                except Exception:
                    pass
            for r in range(25, 29):
                cell = ws_ter.cell(row=r, column=10)
                cell.value = 0.05
                try:
                    cell.number_format = "0%"
                except Exception:
                    pass

        # ------- NEW: build nested save path under Summaries -------
        dept_name = self.get_department_for_teacher(teacher_name) or "UnknownDept"

        base_root = os.path.join(os.path.expanduser("~"), "Documents", "MyWork", "Summaries")
        # Clean pieces to be safe for folders
        safe_dept = _safe_filename(dept_name)
        safe_teacher = _safe_filename(teacher_name)
        safe_ay = _safe_filename(ay)
        safe_sem = _safe_filename(sem)

        out_dir = os.path.join(base_root, safe_dept, safe_ay, safe_sem, safe_teacher)
        os.makedirs(out_dir, exist_ok=True)

        file_title = f"{teacher_name}, {ay}, {sem}"
        filename = _safe_filename(file_title) + ".xlsx"
        save_path = os.path.join(out_dir, filename)
        # -----------------------------------------------------------

        _apply_workbook_protection(wb, password="ATSLock")
        wb.save(save_path)
        return save_path


    # ---------- Internals ----------
    def _open_summary_popup(self, master, teacher_name: str):
        # center window
        sw = master.winfo_screenwidth()
        sh = master.winfo_screenheight()
        x = (sw - 1280) // 2
        y = (sh - 720) // 2

        win = ctk.CTkToplevel(master)
        win.title(f"Teaching Efficiency Rating - {teacher_name}")
        win.geometry(f"1280x720+{x}+{y}")
        win.configure(fg_color="#F3F4F6")
        win.transient(master)
        win.grab_set()
        win.focus_force()
        win.lift()

        scroll = ctk.CTkScrollableFrame(
            win, fg_color="#FFFFFF",
            label_font=("Poppins", 14, "bold"),
            label_text_color="#DC2626"
        )
        scroll.pack(fill="both", expand=True, padx=20, pady=20)

        self.entry_widgets = {}

        def add_row(parent, row_idx, label, percent="", key=None):
            ctk.CTkLabel(parent, text=label, font=("Poppins", 12),
                         text_color="#1F2937", anchor="w").grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
            rating_entry = ctk.CTkEntry(parent, width=120, font=("Poppins", 12))
            rating_entry.grid(row=row_idx, column=1, padx=5, pady=2)
            eq_entry = ctk.CTkEntry(parent, width=150, font=("Poppins", 12))
            eq_entry.grid(row=row_idx, column=2, padx=5, pady=2)
            ctk.CTkLabel(parent, text=percent, font=("Poppins", 12),
                         text_color="#1F2937", anchor="center").grid(row=row_idx, column=3, padx=5, pady=2)
            point_entry = ctk.CTkEntry(parent, width=120, font=("Poppins", 12))
            point_entry.grid(row=row_idx, column=4, padx=5, pady=2)
            if key:
                self.entry_widgets[key] = {"rating": rating_entry, "equivalent": eq_entry, "point": point_entry}

        def add_behavior_row(parent, row_idx, label, percent="", key=None, include_in_calc=True):
            ctk.CTkLabel(parent, text=label, font=("Poppins", 12),
                         text_color="#1F2937", anchor="w").grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
            eq_entry = ctk.CTkEntry(parent, width=150, font=("Poppins", 12))
            eq_entry.grid(row=row_idx, column=2, padx=5, pady=2)
            ctk.CTkLabel(parent, text=percent, font=("Poppins", 12),
                         text_color="#1F2937", anchor="center").grid(row=row_idx, column=3, padx=5, pady=2)
            point_entry = ctk.CTkEntry(parent, width=120, font=("Poppins", 12))
            point_entry.grid(row=row_idx, column=4, padx=5, pady=2)

            if key:
                weight = float(percent.strip('%')) if percent else 0.0
                # store whether this row should participate in the overall calculation
                self.entry_widgets[key] = {
                    "equivalent": eq_entry,
                    "point": point_entry,
                    "weight": weight,
                    "include": bool(include_in_calc),
                }

                # If this is not included, disable the inputs so the user sees it
                if not include_in_calc:
                    eq_entry.configure(state="disabled")
                    point_entry.configure(state="disabled")

                def update_point(_event=None, entry_key=key):
                    # If this row is not included in the calculations, keep point blank
                    if not self.entry_widgets[entry_key].get("include", True):
                        point_entry.delete(0, "end")
                        self._update_overall_point_score()
                        return

                    val = eq_entry.get().strip()
                    try:
                        num = float(val)
                        weight_local = self.entry_widgets[entry_key]["weight"] / 100.0
                        point = num * weight_local
                        point_entry.delete(0, "end")
                        point_entry.insert(0, f"{point:.2f}")
                    except ValueError:
                        point_entry.delete(0, "end")
                    self._update_overall_point_score()

                eq_entry.bind("<KeyRelease>", update_point)


        # Header
        ctk.CTkLabel(scroll, text="TEACHING EFFICIENCY RATING (TER) SCALE FORM",
                     font=("Poppins", 16, "bold"), text_color="#FFFFFF",
                     fg_color="#DC2626", corner_radius=6).pack(fill="x", pady=5)

        # Instructor Info
        info = ctk.CTkFrame(scroll, fg_color="transparent")
        info.pack(fill="x", pady=5)
        for lbl in ["Instructor:", "College:", "Rating Period:", "Department:"]:
            ctk.CTkLabel(info, text=lbl, font=("Poppins", 12), text_color="#1F2937").pack(side="left", padx=5)
            e = ctk.CTkEntry(info, width=150, font=("Poppins", 12))
            e.pack(side="left", padx=10)
            self.entry_widgets[lbl.strip(":").lower()] = e

        # Performance
        perf = ctk.CTkFrame(scroll, fg_color="transparent")
        perf.pack(fill="x", pady=10)
        ctk.CTkLabel(perf, text="I. PERFORMANCE (70%)",
                     font=("Poppins", 14, "bold"), text_color="#DC2626").grid(row=0, column=0, sticky="w", pady=(0,5))
        headers = ["", "RATING", "RATING EQUIVALENT", "RATING %", "POINT SCORE"]
        for i, h in enumerate(headers):
            ctk.CTkLabel(perf, text=h, font=("Poppins", 12, "bold"), text_color="#1F2937").grid(row=1, column=i, padx=5, pady=2)

        ctk.CTkLabel(perf, text="1. Instruction (55%)",
                     font=("Poppins", 12, "bold"), text_color="#DC2626").grid(row=2, column=0, sticky="w", pady=(0,3))
        add_row(perf, 3, "a) Student as Rater", "25%", key="student_rater")
        add_row(perf, 4, "b) Peer as Rater",    "15%", key="peer_rater")
        add_row(perf, 5, "c) Dean as Rater",    "15%", key="dean_rater")

        add_behavior_row(perf, 6, "2. Research",   "5%", key="research")
        add_behavior_row(perf, 7, "3. Extension",  "5%", key="extension")
        add_behavior_row(perf, 8, "4. Production", "5%", key="production")
        
        # Determine if the current teacher is a dean (for behavior rows)
        is_dean = False
        try:
            is_dean = self._is_teacher_dean(teacher_name)
        except Exception:
            is_dean = False


        # Behavior
        beh = ctk.CTkFrame(scroll, fg_color="transparent")
        beh.pack(fill="x", pady=10)
        ctk.CTkLabel(beh, text="II. BEHAVIOR (30%)",
                     font=("Poppins", 14, "bold"), text_color="#DC2626").grid(row=0, column=0, sticky="w", pady=(0,5))
        for i, h in enumerate(headers):
            ctk.CTkLabel(beh, text=h, font=("Poppins", 12, "bold"), text_color="#1F2937").grid(row=1, column=i, padx=5, pady=2)

        # Adjust percentages based on whether teacher is a dean
        if is_dean:
            # When evaluating a dean: all 6 criteria at 5% each = 30%
            add_behavior_row(beh, 2, "1. Courtesy",                      "5%", key="courtesy")
            add_behavior_row(beh, 3, "2. Human Relations",               "5%", key="human_relations")
            add_behavior_row(beh, 4, "3. Punctuality and Attendance",    "5%", key="punctuality")
            add_behavior_row(beh, 5, "4. Initiative",                    "5%", key="initiative")
            add_behavior_row(beh, 6, "5. Leadership (Supervisors only)", "5%", key="leadership", include_in_calc=True)
            add_behavior_row(beh, 7, "6. Stress Tolerance (Supervisors only)", "5%", key="stress_tolerance", include_in_calc=True)
        else:
            # When evaluating regular faculty: 4 criteria at 7.5% each = 30%
            add_behavior_row(beh, 2, "1. Courtesy",                      "7.5%", key="courtesy")
            add_behavior_row(beh, 3, "2. Human Relations",               "7.5%", key="human_relations")
            add_behavior_row(beh, 4, "3. Punctuality and Attendance",    "7.5%", key="punctuality")
            add_behavior_row(beh, 5, "4. Initiative",                    "7.5%", key="initiative")


        # Plus Factor
        plus = ctk.CTkFrame(scroll, fg_color="transparent")
        plus.pack(fill="x", pady=10)
        ctk.CTkLabel(plus, text="PLUS FACTOR (not to exceed one (1) credit point)",
                     font=("Poppins", 12, "bold"), text_color="#DC2626").grid(row=0, column=0, sticky="w", pady=(0,5))
        e_plus = ctk.CTkEntry(plus, width=120, font=("Poppins", 12))
        e_plus.grid(row=0, column=1, padx=5, pady=2)
        self.entry_widgets["plus_factor"] = e_plus
        e_plus.bind("<KeyRelease>", lambda _e: self._update_overall_point_score())

        # Summary Ratings
        summ = ctk.CTkFrame(scroll, fg_color="transparent")
        summ.pack(fill="x", pady=10)
        labels = [
            ("Overall Point Score", "overall_point"),
            ("Equivalent Numerical Rating", "eq_numerical"),
            ("Equivalent Adjective Rating", "eq_adjective"),
        ]
        for idx, (text, key) in enumerate(labels):
            ctk.CTkLabel(summ, text=text, font=("Poppins", 12, "bold"), text_color="#1F2937").grid(row=idx, column=0, sticky="w", padx=5, pady=2)
            e = ctk.CTkEntry(summ, width=120, font=("Poppins", 12))
            e.grid(row=idx, column=1, padx=5, pady=2)
            self.entry_widgets[key] = e

        # Comments
        comments = ctk.CTkFrame(scroll, fg_color="transparent")
        comments.pack(fill="x", pady=10)
        ctk.CTkLabel(comments, text="EMPLOYEE'S COMMENTS/REMARKS",
                     font=("Poppins", 12, "bold"), text_color="#DC2626").pack(anchor="w")
        self.entry_widgets["employee_comments"] = ctk.CTkTextbox(comments, height=50, width=900, font=("Poppins", 12))
        self.entry_widgets["employee_comments"].pack(pady=5)
        ctk.CTkLabel(comments, text="RATER'S COMMENTS/REMARKS",
                     font=("Poppins", 12, "bold"), text_color="#DC2626").pack(anchor="w")
        self.entry_widgets["rater_comments"] = ctk.CTkTextbox(comments, height=50, width=900, font=("Poppins", 12))
        self.entry_widgets["rater_comments"].pack(pady=5)

        # Save/Close Row (now includes Export)
        btn_row = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_row.pack(fill="x", pady=12)

        def _do_export():
            try:
                # ensure the latest totals are reflected
                self._update_overall_point_score()
                save_path = self.export_full_summary("template.xlsx")
                if messagebox and save_path:
                    messagebox.showinfo("Saved", f"✅ Summary exported:\n{save_path}")
            except Exception as e:
                if messagebox:
                    messagebox.showerror("Export Error", str(e))

        ctk.CTkButton(
            btn_row, text="Export Summary",
            fg_color="#691612", hover_color="#8B1D18",
            text_color="#FFFFFF", font=("Poppins", 14, "bold"),
            corner_radius=8, command=_do_export
        ).pack(side="right", padx=8)

        ctk.CTkButton(
            btn_row, text="Close",
            fg_color="#9CA3AF", hover_color="#6B7280",
            text_color="#FFFFFF", command=win.destroy
        ).pack(side="right")

    # ---------- helpers ----------
    @staticmethod
    def _widget_alive(widget) -> bool:
        try:
            return bool(widget) and (not hasattr(widget, "winfo_exists") or widget.winfo_exists())
        except Exception:
            return False

    def _set_value(self, key, value, field=None):
        if key not in self.entry_widgets:
            return
        widget = self.entry_widgets[key]
        if isinstance(widget, dict):
            if field not in widget:
                return
            widget = widget[field]
        if not self._widget_alive(widget):
            return
        if hasattr(widget, "insert") and hasattr(widget, "delete"):
            if "Textbox" in widget.__class__.__name__:
                widget.delete("1.0", "end")
                widget.insert("1.0", value)
            else:
                widget.delete(0, "end")
                widget.insert(0, value)

    def _update_overall_point_score(self):
        # sum points
        total = 0.0
        for key, widgets in self.entry_widgets.items():
            if isinstance(widgets, dict) and "point" in widgets:
                # skip rows explicitly marked as not included in calculation
                if widgets.get("include") is False:
                    continue
                try:
                    total += float(widgets["point"].get().strip() or 0)
                except Exception:
                    pass
        # plus factor
        if "plus_factor" in self.entry_widgets:
            try:
                total += float(self.entry_widgets["plus_factor"].get().strip() or 0)
            except Exception:
                pass

        if "overall_point" in self.entry_widgets:
            e = self.entry_widgets["overall_point"]
            e.delete(0, "end"); e.insert(0, f"{total:.2f}")

        eq_num = compute_equivalent_numerical(total)
        if "eq_numerical" in self.entry_widgets:
            e = self.entry_widgets["eq_numerical"]
            e.delete(0, "end"); e.insert(0, str(eq_num))

        eq_adj = compute_adjective_rating(eq_num)
        if "eq_adjective" in self.entry_widgets:
            e = self.entry_widgets["eq_adjective"]
            e.delete(0, "end"); e.insert(0, eq_adj)

    # ---------- Excel writers ----------
    def _write_manual_results(self, ws):
        mapping = {
            "instructor": "D9",
            "college": "D10",
            "rating period": "J9",
            "department": "J10",

            "research":      {"equivalent": "F19", "point": "L19"},
            "extension":     {"equivalent": "F20", "point": "L20"},
            "production":    {"equivalent": "F21", "point": "L21"},

            "courtesy":         {"equivalent": "H25", "point": "L25"},
            "human_relations":  {"equivalent": "H26", "point": "L26"},
            "punctuality":      {"equivalent": "H27", "point": "L27"},
            "initiative":       {"equivalent": "H28", "point": "L28"},
            "leadership":       {"equivalent": "H29", "point": "L29"},
            "stress_tolerance": {"equivalent": "H30", "point": "L30"},
        }

        for key, widget in self.entry_widgets.items():
            if key not in mapping:
                continue

            target = mapping[key]
            if isinstance(widget, dict):
                for sub_key, entry in widget.items():
                    if sub_key in target:
                        ws[target[sub_key]] = entry.get().strip()
            elif hasattr(widget, "get"):
                if widget.__class__.__name__ == "CTkTextbox":
                    ws[target] = widget.get("1.0", "end").strip()
                else:
                    ws[target] = widget.get().strip()

    def _write_student_scores(self, ws):
        docs = self.processed_results.get(self.current_teacher, {}).get("Student", [])
        if not docs:
            return
        row_maps = {
            "Section 1": {1: 12, 2: 13, 3: 14, 4: 15, 5: 16},
            "Section 2": {1: 18, 2: 19, 3: 20, 4: 21, 5: 22},
            "Section 3": {1: 24, 2: 25, 3: 26, 4: 27, 5: 28},
            "Section 4": {1: 30, 2: 31, 3: 32, 4: 33, 5: 34},
        }
        start_col = 6  # F
        for doc_idx, (_name, result, *_) in enumerate(docs):
            col = start_col + doc_idx
            for section, row_map in row_maps.items():
                for rownum, score in result.get(section, {}).items():
                    excel_row = row_map.get(rownum)
                    if excel_row:
                        ws.cell(row=excel_row, column=col, value=score)

    def _write_peer_scores(self, ws):
        docs = self.processed_results.get(self.current_teacher, {}).get("Peer", [])
        if not docs:
            return
        row_maps = {
            "Section 1": {1: 12, 2: 13, 3: 14, 4: 15, 5: 16},
            "Section 2": {1: 18, 2: 19, 3: 20, 4: 21, 5: 22},
            "Section 3": {1: 24, 2: 25, 3: 26, 4: 27, 5: 28},
            "Section 4": {1: 30, 2: 31, 3: 32, 4: 33, 5: 34},
        }
        start_col = 345  # MG
        for doc_idx, (_name, result, *_) in enumerate(docs):
            col = start_col + doc_idx
            for section, row_map in row_maps.items():
                for rownum, score in result.get(section, {}).items():
                    excel_row = row_map.get(rownum)
                    if excel_row:
                        ws.cell(row=excel_row, column=col, value=score)

    def _write_self_scores(self, ws):
        docs = self.processed_results.get(self.current_teacher, {}).get("Self", [])
        if not docs:
            return
        _name, result, *_ = docs[0]
        row_maps = {
            "Section 1": {1: 46, 2: 47, 3: 48, 4: 49, 5: 50},
            "Section 2": {1: 52, 2: 53, 3: 54, 4: 55, 5: 56},
            "Section 3": {1: 58, 2: 59, 3: 60, 4: 61, 5: 62},
            "Section 4": {1: 64, 2: 65, 3: 66, 4: 67, 5: 68},
        }
        col = 3  # C
        for section, row_map in row_maps.items():
            for rownum, score in result.get(section, {}).items():
                excel_row = row_map.get(rownum)
                if excel_row:
                    ws.cell(row=excel_row, column=col, value=score)

    def _write_dean_scores(self, ws):
        docs = self.processed_results.get(self.current_teacher, {}).get("Dean", [])
        if not docs:
            return
        _name, result, *_ = docs[0]
        row_maps = {
            "Section 1": {1: 46, 2: 47, 3: 48, 4: 49, 5: 50},
            "Section 2": {1: 52, 2: 53, 3: 54, 4: 55, 5: 56},
            "Section 3": {1: 58, 2: 59, 3: 60, 4: 61, 5: 62},
            "Section 4": {1: 64, 2: 65, 3: 66, 4: 67, 5: 68},
        }
        col = 15  # O
        for section, row_map in row_maps.items():
            for rownum, score in result.get(section, {}).items():
                excel_row = row_map.get(rownum)
                if excel_row:
                    ws.cell(row=excel_row, column=col, value=score)

    # -----------------other helpers-------------------------
    def _is_teacher_dean(self, teacher_name: str) -> bool:
        """Return True if the given teacher is set as a department dean."""
        if not self.db or not teacher_name:
            return False
        try:
            with self.db.connect() as conn:
                row = conn.execute("""
                    SELECT 1
                    FROM departments d
                    JOIN faculty f ON f.id = d.dean_id
                    WHERE f.full_name = ?
                """, (teacher_name,)).fetchone()
            return row is not None
        except Exception:
            return False

    def _get_teacher_meta(self, teacher_name: str):
        """Return dict with name parts, rank, subrank, department, dean name."""
        meta = {
            "first_name": "",
            "middle_name": "",
            "last_name": "",
            "rank": "",
            "subrank": "",
            "department": "",
            "dean_name": "",
            "dean_title": "Dean",
        }
        if not self.db or not teacher_name:
            return meta

        def _split_rank(role_text: str):
            """Best-effort split of role text into (rank, subrank)."""
            if not role_text:
                return "", ""
            role_text = role_text.strip()
            rank_labels = [
                "Assistant Professor",
                "Associate Professor",
                "Professor",
                "Instructor",
            ]
            roman = {"I", "II", "III", "IV", "V", "VI"}
            for label in rank_labels:
                if role_text.startswith(label):
                    tail = role_text[len(label):].strip()
                    sub = tail if tail in roman else (tail or "")
                    return label, sub
            # Fallback: treat whole string as rank
            return role_text, ""

        try:
            with self.db.connect() as conn:
                row = conn.execute(
                    """
                    SELECT f.first_name, f.middle_name, f.last_name, f.role,
                           d.name AS department,
                           dean.full_name AS dean_name
                    FROM faculty f
                    JOIN departments d ON d.id = f.department_id
                    LEFT JOIN faculty dean ON d.dean_id = dean.id
                    WHERE f.full_name = ?
                    """,
                    (teacher_name,),
                ).fetchone()
            if row:
                meta["first_name"] = row[0] or ""
                meta["middle_name"] = row[1] or ""
                meta["last_name"] = row[2] or ""
                rank_label, sub_label = _split_rank(row[3] or "")
                meta["rank"] = rank_label
                meta["subrank"] = sub_label
                meta["department"] = row[4] or ""
                meta["dean_name"] = row[5] or ""
        except Exception:
            pass
        return meta

    
