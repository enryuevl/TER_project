from customtkinter import *
from PIL import Image, ImageTk
from customtkinter import CTkImage
import pickle
import os
from tkinter import messagebox, filedialog
import main_code
import pandas as pd
import cv2, numpy as np, os, threading
from scanner import WIAScanner
import db
import pythoncom
from utils import set_sidebar_state
import utils
from summary_helpers import protect_workbook_file
import datetime
import shutil
import json

class ScanPage:
    def __init__(self, master, processed_results):
        """Initialize the Scan Page inside a given frame."""
        self.master = master
        self.processed_results = processed_results
        self.last_processed_times = {}
        self.annotated_cache = {}  # Cache for annotated images
        self.current_scan_dir = None
        self.scanner = None
        self.load_results()
        # call summary controller
        from summary_helpers import SummaryFormController
        self.summary = SummaryFormController(self.processed_results, db_module=db)


        # Tkinter variables
        self.teacher_var = StringVar()
        self.subject_var = StringVar()
        self.block_var = StringVar()

        # ID mappings
        self.teacher_name_to_id = {}
        self.subject_code_to_id = {}
        self.block_label_to_id = {}
        self.subject_label_to_id = {}   
        self.subject_code_by_label = {}

        # Build UI
        self._build_ui()

        # Load teacher list from DB
        self.load_teachers()
        
    def _show_validation_dialog(self, title: str, message: str, badge_text: str = "NOTICE"):
        """
        ATS-themed simple validation dialog with a single OK button.
        Used instead of messagebox.showwarning for things like 'Select Subject'.
        """
        TOPBAR = "#BF3131"
        PANEL_BG = "#F5F5F5"
        LIGHT_TEXT = "#FFEFEF"
        WHITE = "#FFFFFF"

        parent = self.master

        dialog = CTkToplevel(parent)
        dialog.title(title)
        dialog.resizable(False, False)

        dialog.transient(parent)
        dialog.grab_set()

        main_frame = CTkFrame(dialog, fg_color=PANEL_BG, corner_radius=12)
        main_frame.pack(fill="both", expand=True, padx=16, pady=16)

        # Header
        header = CTkFrame(main_frame, fg_color=TOPBAR, corner_radius=10)
        header.pack(fill="x", padx=8, pady=(8, 4))

        CTkLabel(
            header,
            text=title,
            font=("Poppins", 16, "bold"),
            text_color=WHITE
        ).pack(side="left", padx=12, pady=8)

        CTkLabel(
            header,
            text=badge_text.upper(),
            font=("Poppins", 11, "bold"),
            text_color=TOPBAR,
            fg_color=LIGHT_TEXT,
            corner_radius=999,
            padx=10,
            pady=4,
        ).pack(side="right", padx=12, pady=8)

        # Body
        body = CTkFrame(main_frame, fg_color=WHITE, corner_radius=10)
        body.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        CTkLabel(
            body,
            text=message,
            font=("Poppins", 12),
            text_color="#333333",
            justify="left",
        ).pack(anchor="w", padx=12, pady=(12, 8))

        # Footer
        footer = CTkFrame(main_frame, fg_color=PANEL_BG)
        footer.pack(fill="x", padx=8, pady=(0, 8))

        def close_dialog():
            dialog.destroy()

        CTkButton(
            footer,
            text="OK",
            font=("Poppins", 12, "bold"),
            fg_color="#AC5353",
            hover_color="#BF3131",
            text_color=WHITE,
            corner_radius=8,
            width=100,
            command=close_dialog
        ).pack(side="right", padx=8, pady=4)

        # Center on parent
        dialog.update_idletasks()
        w = dialog.winfo_width()
        h = dialog.winfo_height()
        if w < 380:
            w = 380
        if h < 180:
            h = 180
        dialog.minsize(w, h)

        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (w // 2)
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (h // 2)
        dialog.geometry(f"{w}x{h}+{x}+{y}")

        dialog.wait_window(dialog)

        
        
    def _open_scan_progress_dialog(self):
        """
        ATS-themed non-closable 'Scanning in progress' dialog.
        Stored as self.scan_dialog so the worker thread can close it.
        """
        TOPBAR = "#BF3131"
        PANEL_BG = "#F5F5F5"
        LIGHT_TEXT = "#FFEFEF"
        WHITE = "#FFFFFF"

        parent = self.master

        dialog = CTkToplevel(parent)
        dialog.title("Scanning in progress")
        dialog.resizable(False, False)

        # Make it modal & prevent closing via [X]
        dialog.transient(parent)
        dialog.grab_set()
        dialog.protocol("WM_DELETE_WINDOW", lambda: None)

        main_frame = CTkFrame(dialog, fg_color=PANEL_BG, corner_radius=12)
        main_frame.pack(fill="both", expand=True, padx=16, pady=16)

        # Header bar
        header = CTkFrame(main_frame, fg_color=TOPBAR, corner_radius=10)
        header.pack(fill="x", padx=8, pady=(8, 4))

        CTkLabel(
            header,
            text="Scanning in progress",
            font=("Poppins", 16, "bold"),
            text_color=WHITE
        ).pack(side="left", padx=12, pady=8)

        CTkLabel(
            header,
            text="SCANNER",
            font=("Poppins", 11, "bold"),
            text_color=TOPBAR,
            fg_color=LIGHT_TEXT,
            corner_radius=999,
            padx=10,
            pady=4,
        ).pack(side="right", padx=12, pady=8)

        # Body
        body = CTkFrame(main_frame, fg_color=WHITE, corner_radius=10)
        body.pack(fill="both", expand=False, padx=8, pady=(4, 8))

        CTkLabel(
            body,
            text=(
                "Please wait while the documents are being scanned and processed.\n"
                "This may take a few moments depending on the number of pages.\n\n"
                "Do not close the application during this process."
            ),
            font=("Poppins", 12),
            text_color="#333333",
            justify="left",
        ).pack(anchor="w", padx=12, pady=(12, 8))

        # Progress bar (fake/indeterminate style)
        bar_frame = CTkFrame(body, fg_color="transparent")
        bar_frame.pack(fill="x", padx=12, pady=(0, 12))

        progress = CTkProgressBar(bar_frame, height=10, corner_radius=5)
        progress.pack(fill="x")
        progress.set(0.3)  # static, just to give visual feedback

        # Footer
        footer = CTkFrame(main_frame, fg_color=PANEL_BG)
        footer.pack(fill="x", padx=8, pady=(0, 8))

        CTkLabel(
            footer,
            text="Scanning… please wait",
            font=("Poppins", 11, "italic"),
            text_color="#6B7280",
        ).pack(side="right", padx=8, pady=4)

        # Center dialog over parent
        dialog.update_idletasks()
        w = dialog.winfo_width()
        h = dialog.winfo_height()
        if w < 420:
            w = 420
        if h < 220:
            h = 220
        dialog.minsize(w, h)

        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (w // 2)
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (h // 2)
        dialog.geometry(f"{w}x{h}+{x}+{y}")

        # keep reference so we can close later
        self.scan_dialog = dialog



    # ---------------- UI BUILDERS ---------------- #
    def _build_ui(self):
        """Create the layout and widgets."""
        for widget in self.master.winfo_children():
            widget.destroy()
            
        
        # ⬅️ Add red header bar
        title_bar = CTkFrame(self.master, fg_color="#BF3131", height=70, corner_radius=10)
        title_bar.pack(fill="x", padx=20, pady=(10, 15))
        CTkLabel(
            title_bar,
            text="Scan & Preview",
            font=("Montserrat", 20, "bold"),
            text_color="#FFFFFF"
        ).pack(side="left", padx=25, pady=12)


        self.content_frame = CTkFrame(self.master, fg_color="#F8F9FA")
        self.content_frame.pack(fill="both", expand=True, padx=20, pady=20)

        main_layout = CTkFrame(self.content_frame, fg_color="transparent")
        main_layout.pack(fill="both", expand=True, padx=10)
        main_layout.grid_columnconfigure(0, weight=1)
        main_layout.grid_columnconfigure(1, weight=3)

        # Left column
        left_column = CTkFrame(main_layout, fg_color="transparent")
        left_column.grid(row=0, column=0, sticky="nsew", padx=10)

        self._build_scanner_controls(left_column)
        self._build_teacher_dropdown(left_column)
        self._build_results_panel(left_column)

        # Right column (Preview)
        right_column = CTkFrame(main_layout, fg_color="transparent")
        right_column.grid(row=0, column=1, sticky="nsew", padx=10)
        self._build_preview_panel(right_column)
        
    def show_qc_error_dialog(self, qc_errors):
        """
        Custom QC Error Dialog that matches the ATS palette.
        qc_errors: list of (filename, reason) tuples.
        """
        TOPBAR = "#BF3131"     
        SIDEBAR_BTN = "#AC5353" 
        HOVER = "#BF3131"      
        PANEL_BG = "#F5F5F5"    
        LIGHT_TEXT = "#FFEFEF"
        WHITE = "#FFFFFF"

        parent = self.master  # parent window for the dialog

        # --- Dialog Window ---
        dialog = CTkToplevel(parent)
        dialog.title("QC Errors Detected")
        dialog.geometry("560x360")
        dialog.resizable(False, False)

        dialog.transient(parent)
        dialog.grab_set()   # modal

        # --- Main container ---
        main_frame = CTkFrame(dialog, fg_color=PANEL_BG, corner_radius=12)
        main_frame.pack(fill="both", expand=True, padx=16, pady=16)

        # --- Header bar ---
        header = CTkFrame(main_frame, fg_color=TOPBAR, corner_radius=10)
        header.pack(fill="x", padx=8, pady=(8, 4))

        CTkLabel(
            header,
            text="Incomplete / Blank Pages Detected",
            font=("Poppins", 16, "bold"),
            text_color=WHITE
        ).pack(side="left", padx=12, pady=8)

        CTkLabel(
            header,
            text="ERROR",
            font=("Poppins", 11, "bold"),
            text_color=TOPBAR,
            fg_color=LIGHT_TEXT,
            corner_radius=999,
            padx=10,
            pady=4
        ).pack(side="right", padx=12, pady=8)

        # --- Body ---
        body = CTkFrame(main_frame, fg_color=WHITE, corner_radius=10)
        body.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        CTkLabel(
            body,
            text=(
                "The following documents have missing keys and were discarded.\n"
                "Please rescan these page(s):"
            ),
            font=("Poppins", 12),
            text_color="#333333",
            justify="left"
        ).pack(anchor="w", padx=12, pady=(10, 6))

        # --- Scrollable list of errors ---
        list_frame = CTkScrollableFrame(
            body,
            fg_color=PANEL_BG,
            corner_radius=8,
            height=160
        )
        list_frame.pack(fill="both", expand=True, padx=12, pady=(0, 10))

        if qc_errors:
            for fname, reason in qc_errors:
                CTkLabel(
                    list_frame,
                    text=f"• {fname} → {reason}",
                    font=("Poppins", 11),
                    text_color="#333333",
                    anchor="w",
                    justify="left"
                ).pack(fill="x", pady=2)
        else:
            CTkLabel(
                list_frame,
                text="No details available.",
                font=("Poppins", 11, "italic"),
                text_color="#555555"
            ).pack(pady=10)

        # --- Footer buttons ---
        footer = CTkFrame(main_frame, fg_color=PANEL_BG)
        footer.pack(fill="x", padx=8, pady=(0, 8))

        def close_dialog():
            dialog.destroy()

        CTkButton(
            footer,
            text="OK, I will rescan",
            font=("Poppins", 12, "bold"),
            fg_color=SIDEBAR_BTN,
            hover_color=HOVER,
            text_color=WHITE,
            corner_radius=8,
            width=150,
            command=close_dialog
        ).pack(side="right", padx=12, pady=4)

        # Center relative to parent
        dialog.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (dialog.winfo_width() // 2)
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")

        dialog.wait_window(dialog)

        
    def _build_scanner_controls(self, parent):
        scanner_frame = CTkFrame(parent, fg_color="#FFFFFF", corner_radius=10)
        scanner_frame.pack(fill="x", pady=10)

        CTkLabel(
            scanner_frame,
            text="Document Scanner",
            font=("Montserrat", 18, "bold"),
            text_color="#334155"
        ).pack(pady=(15, 10), padx=15, anchor="w")

        # Container for dropdown + scan button
        controls_frame = CTkFrame(scanner_frame, fg_color="transparent")
        controls_frame.pack(fill="x", pady=10, padx=15)

        # --- Scanner dropdown label ---
        CTkLabel(
            controls_frame,
            text="Select Scanner",
            font=("Montserrat", 16, "bold"),
            text_color="#334155"
        ).pack(anchor="w", pady=(0, 4))

        # Try to get available scanner names (adjust list_devices() to your WIAScanner API)
        devices = []
        try:
            self.scanner = self.scanner or WIAScanner()
            if hasattr(self.scanner, "list_devices"):
                devices = self.scanner.list_devices() or []
        except Exception:
            devices = []

        if not devices:
            devices = ["No scanners detected"]
            state = "disabled"
        else:
            state = "normal"

        # --- Scanner dropdown ---
        self.scanner_var = StringVar(value=devices[0])
        self.scanner_dropdown = CTkOptionMenu(
            controls_frame,
            variable=self.scanner_var,
            values=devices,
            width=250,
            height=35,
            font=("Montserrat", 14),
            state=state,
            fg_color="#BF3131",
            button_color="#691612",
            button_hover_color="#8E1616",
            text_color="#FFFFFF",
            dropdown_fg_color="#FFFFFF",
            dropdown_hover_color="#F3D0D0",
            dropdown_text_color="#333333",
            dropdown_font=("Montserrat", 14),
        )
        self.scanner_dropdown.pack(fill="x", pady=(0, 10))

        # --- Scan button (kept) ---
        self.btn_scan = CTkButton(
            controls_frame,
            text="Scan",
            command=self.start_scan,
            fg_color="#691612",
            hover_color="#550d0a",
            text_color="#FFFFFF",
            font=("Montserrat", 14),
            height=40,
            corner_radius=8
        )
        self.btn_scan.pack(fill="x", pady=5)

        # Status text
        self.status_label = CTkLabel(
            scanner_frame,
            text="Scanner disconnected" if state == "disabled" else "Scanner ready",
            font=("Montserrat", 14),
            text_color="#64748B"
        )
        self.status_label.pack(pady=5, padx=15, anchor="w")
   

    def _build_teacher_dropdown(self, parent):
        teacher_frame = CTkFrame(parent, fg_color="#FFFFFF", corner_radius=10)
        teacher_frame.pack(fill="x", pady=10)

        # --- Teacher ---
        CTkLabel(teacher_frame, text="Select Teacher",
                font=('Montserrat', 16, 'bold'), text_color="#334155"
        ).pack(pady=(10, 5), padx=15, anchor="w")

        self.teacher_dropdown = CTkOptionMenu(
            teacher_frame, variable=self.teacher_var, values=["Loading..."],
            width=250, height=35, font=('Montserrat', 14),
            fg_color="#BF3131", button_color="#691612", button_hover_color="#8E1616",
            text_color="#FFFFFF",
            dropdown_fg_color="#FFFFFF", dropdown_hover_color="#F3D0D0",
            dropdown_text_color="#333333", dropdown_font=('Montserrat', 14),
        )
        self.teacher_dropdown.pack(padx=15, pady=(10, 5))

        # --- Subject (auto-populated after picking teacher) ---
        CTkLabel(teacher_frame, text="Select Subject",
                font=('Montserrat', 16, 'bold'), text_color="#334155"
        ).pack(pady=(12, 5), padx=15, anchor="w")

        self.subject_dropdown = CTkOptionMenu(
            teacher_frame, variable=self.subject_var,
            values=["— Select a teacher first —"],
            width=250, height=35, font=('Montserrat', 14),
            fg_color="#BF3131", button_color="#691612", button_hover_color="#8E1616",
            text_color="#FFFFFF",
            dropdown_fg_color="#FFFFFF", dropdown_hover_color="#F3D0D0",
            dropdown_text_color="#333333", dropdown_font=('Montserrat', 14),
            state="disabled",
        )
        self.subject_dropdown.pack(padx=15, pady=(10, 5))

        # --- Rater ---
        CTkLabel(teacher_frame, text="Select Rater Type",
                font=('Montserrat', 16, 'bold'), text_color="#334155"
        ).pack(pady=(12, 5), padx=15, anchor="w")

        self.rater_var = StringVar(value="Student")
        self.rater_dropdown = CTkOptionMenu(
            teacher_frame, variable=self.rater_var, values=["Student", "Peer", "Self"],
            width=250, height=35, font=('Montserrat', 14),
            fg_color="#BF3131", button_color="#691612", button_hover_color="#8E1616",
            text_color="#FFFFFF",
            dropdown_fg_color="#FFFFFF", dropdown_hover_color="#F3D0D0",
            dropdown_text_color="#333333", dropdown_font=('Montserrat', 14),
        )
        self.rater_dropdown.pack(padx=15, pady=(10, 5))

        # Bindings
        self.teacher_dropdown.configure(command=lambda *_: self._on_teacher_change())
        self.subject_dropdown.configure(command=lambda *_: self._refresh_teacher_progress())



    def _build_results_panel(self, parent):
        results_frame = CTkFrame(parent, fg_color="#FFFFFF", corner_radius=10)
        results_frame.pack(fill="x", pady=10)

        CTkLabel(results_frame, text="Evaluation Results",
                 font=('Montserrat', 18, 'bold'),
                 text_color="#334155").pack(pady=(15, 10), padx=15, anchor="w")

        self.progress_bar = CTkProgressBar(results_frame, width=250, height=10,
                                           corner_radius=5, progress_color="#BF3131")
        self.progress_bar.pack(fill="x", padx=15, pady=5)
        self.progress_bar.set(0)

        self.scan_info_label = CTkLabel(results_frame, text="No scan loaded",
                                        font=('Montserrat', 12), text_color="#64748B")
        self.scan_info_label.pack(anchor="w", padx=15, pady=5)


    def _build_preview_panel(self, parent):
        preview_frame = CTkFrame(parent, fg_color="#FFFFFF", corner_radius=10)
        preview_frame.pack(fill="both", expand=True)

        CTkLabel(preview_frame, text="Document Preview",
                font=('Montserrat', 18, 'bold'), text_color="#334155").pack(pady=(15, 0))

                # --- Preview file chooser ---
        self.document_listbox = CTkOptionMenu(
            preview_frame,
            values=["No documents loaded"],
            width=250, height=35, font=('Montserrat', 14),

            fg_color="#BF3131",
            button_color="#691612",
            button_hover_color="#8E1616",
            text_color="#FFFFFF",

            dropdown_fg_color="#FFFFFF",
            dropdown_hover_color="#F3D0D0",
            dropdown_text_color="#333333",
            dropdown_font=('Montserrat', 14),
        )
        self.document_listbox.pack(padx=20, pady=10, fill="x")

        # ⬇️ Restore the preview label (this is what was missing)
        self.img_label = CTkLabel(
            preview_frame, text="No image loaded",
            font=('Montserrat', 14), fg_color="#555555",
            width=400, height=500
        )
        self.img_label.pack(padx=10, pady=20, fill="both", expand=True)

        # When user picks a file, update the preview
        self.document_listbox.configure(
            command=lambda choice: self.display_image(
                choice, self.teacher_var.get(), base_dir=self.current_scan_dir
            )
        )



    # ---------------- DB HANDLERS ---------------- #
    def load_teachers(self):
        try:
            with db.connect() as conn:
                rows = conn.execute(
                    "SELECT id, full_name FROM faculty ORDER BY full_name"
                ).fetchall()

            if rows:
                # Map teacher full_name → faculty_id
                self.teacher_name_to_id = {full_name: fid for fid, full_name in rows}
                names = list(self.teacher_name_to_id.keys())

                self.teacher_dropdown.configure(values=names)
                self.teacher_dropdown.set(names[0])
            else:
                self.teacher_dropdown.configure(values=["No teachers found"])

        except Exception as e:
            messagebox.showerror("DB Error", str(e))

    #
    def _infer_ay_and_sem_from_today(self):
        now = datetime.datetime.now()
        y, m = now.year, now.month
        if 8 <= m <= 12:
            return f"{y}-{y+1}", "1st"
        elif 1 <= m <= 6:
            return f"{y-1}-{y}", "2nd"
        else:
            return f"{y-1}-{y}", "Summer"

    def _get_teacher_department(self, teacher_full_name: str) -> str | None:
        try:
            with db.connect() as conn:
                row = conn.execute("""
                    SELECT d.name
                    FROM faculty f
                    JOIN departments d ON d.id = f.department_id
                    WHERE f.full_name = ?
                """, (teacher_full_name,)).fetchone()
            return row[0] if row else None
        except Exception:
            return None

    def _build_scan_dir(self, teacher_full_name: str) -> str:
        dept = self._get_teacher_department(teacher_full_name) or "UnknownDept"
        ay, sem = self._infer_ay_and_sem_from_today()
        folder = os.path.join(os.path.expanduser("~"), "Documents", "MyWork", "Scan", dept, ay, sem, teacher_full_name)
        os.makedirs(folder, exist_ok=True)
        return folder



    # ---------------- SCANNER ACTIONS ---------------- #
    def start_scan(self):
        # -------- fast, synchronous validations (before disabling UI) --------
        teacher = (self.teacher_var.get() or "").strip()
        if not teacher or teacher in ("Loading...", "No teachers found"):
            self._show_validation_dialog(
                "Select Teacher",
                "Please select a teacher before starting the scan.",
                badge_text="Warning"
            )
            return

        label = self.subject_var.get() if hasattr(self, "subject_var") else ""
        subj_code = self.subject_code_by_label.get(label, "")
        if not subj_code:
            self._show_validation_dialog(
                "Select Subject",
                "Please select a subject for this scan.",
                badge_text="Warning"
            )
            return

        # -------- disable controls + show custom 'scanning in progress' dialog --------
        self.set_controls_state("disabled")
        set_sidebar_state("disabled")

        self._open_scan_progress_dialog()

        # -------- worker thread --------
        def worker():
            try:
                pythoncom.CoInitialize()

                # cache current selections for the thread
                rater = self.rater_var.get() if hasattr(self, "rater_var") else "Student"
                selected_scanner = self.scanner_var.get() if hasattr(self, "scanner_var") else None

                self.current_scan_dir = self._build_scan_dir(teacher)

                # activity log
                user = utils.get_current_user()
                db.log_activity(
                    action="scan_started",
                    actor_name=user.get("name"),
                    actor_role=user.get("role"),
                    department_id=user.get("department_id"),
                    teacher_name=teacher,
                    rater_type=rater,
                    details={"subject_code": subj_code}
                )

                # scan
                scanner = WIAScanner(teacher_name=teacher, output_dir=self.current_scan_dir)
                info = scanner.initialize(device_name=selected_scanner)
                self.status_label.configure(text=f"Scanner detected: {info['name']}")
                scanner.create_batch_dir()
                pages, batch_dir = scanner.scan_batch()

                if pages > 0:
                    results, qc_errors = self.process_work_folder(teacher, src_folder=batch_dir)

                    if qc_errors:
                        self.show_qc_error_dialog(qc_errors)

                    if results:
                        self.processed_results.update(results)
                        self.save_results()
                        self._refresh_teacher_progress()
                        try:
                            self._auto_export_summary(teacher)
                        except Exception as ex:
                            messagebox.showerror("Excel Export Error", str(ex))

                        teacher_files = results.get(teacher, {}).get(rater, [])
                        if teacher_files:
                            self.update_preview(teacher_files)

                    db.log_activity(
                        action="scan_completed",
                        actor_name=user.get("name"),
                        actor_role=user.get("role"),
                        department_id=user.get("department_id"),
                        teacher_name=teacher,
                        rater_type=rater,
                        details={"pages_scanned": int(pages), "subject_code": subj_code}
                    )

                    self.status_label.configure(
                        text="Processing complete!" if (results or qc_errors) else "No new documents found."
                    )
                else:
                    self.status_label.configure(text="No documents found.")

            except Exception as e:
                messagebox.showerror("Scan Error", str(e))
            finally:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

                self.set_controls_state("normal")
                set_sidebar_state("normal")

                # close custom scan dialog
                try:
                    if hasattr(self, "scan_dialog") and self.scan_dialog.winfo_exists():
                        self.scan_dialog.destroy()
                except Exception:
                    pass

        threading.Thread(target=worker, daemon=True).start()




    def scan_existing(self):
        teacher = (self.teacher_var.get() or "").strip()
        if not teacher or teacher in ("Loading...", "No teachers found"):
            messagebox.showwarning("No Teacher", "Please select a teacher.")
            return

        scan_dir = self._build_scan_dir(teacher)
        self.current_scan_dir = scan_dir

        results, qc_errors = self.process_work_folder(teacher, base_dir=scan_dir)

        if qc_errors:
            self.show_qc_error_dialog(qc_errors)

        if results:
            # merge in-memory
            self.processed_results.update(results)

            # persist to PKL (optional but recommended for consistency)
            try:
                self.save_results()
            except Exception as e:
                messagebox.showerror("Save Error", str(e))

            # refresh preview & progress
            rater = self.rater_var.get() if hasattr(self, "rater_var") else "Unknown"
            self.update_preview(results.get(teacher, {}).get(rater, []))
            self._refresh_teacher_progress()

            # export Excel only when there were new results
            try:
                self._auto_export_summary(teacher)
            except Exception as ex:
                messagebox.showerror("Excel Export Error", str(ex))
        else:
            # nothing new found; keep UI consistent
            self._refresh_teacher_progress()

    def clear_scan(self):
        self.img_label.configure(image=None, text="No image loaded")
        self.document_listbox.configure(values=["No documents loaded"])
        self.processed_results.clear()

    # ---------------- PROCESSING---------------- #
    def _qc_check_page(self, result_dict):
        """
        Returns (is_ok, page_blank, missing_map, total_detected)
        - is_ok = True only if all 4 sections have complete rows (1..5)
        - page_blank = True if nothing detected
        - missing_map = {section: [missing rows]}
        """
        expected = {"Section 1": 5, "Section 2": 5, "Section 3": 5, "Section 4": 5}
        missing_map = {}
        total_detected = 0

        for sec, n in expected.items():
            got = set(result_dict.get(sec, {}).keys())
            total_detected += len(got)
            missing = sorted(set(range(1, n + 1)) - got)
            missing_map[sec] = missing

        page_blank = (total_detected == 0)
        # A page is OK only if no section has missing rows
        is_ok = (not page_blank) and all(len(m) == 0 for m in missing_map.values())
        return is_ok, page_blank, missing_map, total_detected

    def process_scan(self):
        if not self.processed_results:
            messagebox.showwarning("Warning", "No scan found.")
            return
        self._refresh_teacher_progress()
        messagebox.showinfo("Done", "Evaluation processed successfully!")

    def process_work_folder(self, teacher, src_folder=None, base_dir=None):
        
        import time

        # final save folder = Dept/AY/Sem/Teacher
        teacher_root = base_dir or self._build_scan_dir(teacher)
        os.makedirs(teacher_root, exist_ok=True)

        # prefer explicit batch folder
        if src_folder is None:
            src_folder = teacher_root
        os.makedirs(src_folder, exist_ok=True)

        counter_file = os.path.join(teacher_root, "saved_counter.txt")

        def _load_saved_counter():
            try:
                if os.path.exists(counter_file):
                    with open(counter_file, "r") as f:
                        return int(f.read().strip())
            except Exception:
                pass
            return 1

        def _store_saved_counter(n: int):
            try:
                with open(counter_file, "w") as f:
                    f.write(str(n))
            except Exception:
                pass

        save_num = _load_saved_counter()
        new_results, qc_errors = [], []
        rater = self.rater_var.get() if hasattr(self, "rater_var") else "Unknown"

        # --- process files in stable order
        for entry in sorted(os.scandir(src_folder), key=lambda e: e.name.lower()):
            if not entry.name.lower().endswith(".bmp"):
                continue

            img_path = os.path.join(src_folder, entry.name)
            if not os.path.exists(img_path):
                qc_errors.append((entry.name, f"source missing at {img_path}"))
                continue

            # short wait in case Windows still holds the handle
            for _ in range(10):
                try:
                    with open(img_path, "rb"):
                        pass
                    break
                except Exception:
                    time.sleep(0.05)

            img = cv2.imread(img_path)
            if img is None:
                qc_errors.append((entry.name, "cannot read image (locked/corrupt)"))
                continue

            try:
                result_dict, annotated_img = main_code.process_sections(img)
            except Exception as e:
                qc_errors.append((entry.name, f"processing error: {e}"))
                continue

            # 🔒 STRICT QC (your rule): all 4 sections complete; rejects are discarded
            is_ok, page_blank, missing_map, total_detected = self._qc_check_page(result_dict)
            if not is_ok:
                if page_blank:
                    reason = "no marks detected"
                else:
                    parts = [f"{sec} missing {','.join(map(str, miss))}"
                            for sec, miss in missing_map.items() if miss]
                    reason = "incomplete (" + "; ".join(parts) + ")"
                qc_errors.append((entry.name, reason))
                # optional: remove the bad raw scan so it doesn’t linger in _incoming
                try:
                    os.remove(img_path)
                except Exception:
                    pass
                continue  # ❌ do not move, do not consume a save number

            # ✅ ACCEPTED → assign next gapless name in teacher_root
            final_name = f"{save_num}.bmp"                 # or f"{save_num:03d}.bmp"
            final_path = os.path.join(teacher_root, final_name)
            os.makedirs(os.path.dirname(final_path), exist_ok=True)

            try:
                shutil.move(img_path, final_path)
                current_label = self.subject_var.get() if hasattr(self, "subject_var") else ""
                current_code  = self.subject_code_by_label.get(current_label, "")
                self._save_subject_meta(teacher_root, final_name, current_code)
                
                # Add metadata to result_dict for dashboard tracking
                ay, sem = self._infer_ay_and_sem_from_today()
                result_dict["subject_code"] = current_code
                result_dict["academic_year"] = ay
                result_dict["semester"] = sem

            except Exception as e:
                exists_src = os.path.exists(img_path)
                qc_errors.append((entry.name, f"move failed to {final_name}: {e} (src_exists={exists_src})"))
                continue

            # cache annotated preview with FULL PATH key
            self.annotated_cache[final_path] = annotated_img
            new_results.append((final_name, result_dict))
            save_num += 1

        _store_saved_counter(save_num)

        # cleanup empty batch dir (optional)
        try:
            if src_folder != teacher_root and not any(os.scandir(src_folder)):
                os.rmdir(src_folder)
        except Exception:
            pass

        # Only one "Self" page allowed (keep first that PASSED QC)
        if rater == "Self" and len(new_results) > 1:
            keep = new_results[:1]
            for drop_name, _ in new_results[1:]:
                try:
                    os.remove(os.path.join(teacher_root, drop_name))
                except Exception:
                    pass
            new_results = keep
            # (we don't roll back the counter)

        results_dict = {teacher: {rater: new_results}} if new_results else {}
        return results_dict, qc_errors

    # ---------------- RESULT HANDLERS ---------------- #

    def save_results(self, path=None):
        try:
            # 🔹 default path = same folder as database
            if path is None:
                db_path = db.get_default_db_path()
                base_dir = os.path.dirname(db_path)
                path = os.path.join(base_dir, "results.pkl")

            if os.path.exists(path):
                with open(path, "rb") as f:
                    old_data = pickle.load(f)
                old_results = old_data.get("results", {})
                old_times = old_data.get("last_processed_times", {})
            else:
                old_results, old_times = {}, {}

            # --- merge new results ---
            for teacher, rater_dict in self.processed_results.items():
                # 🔑 Normalize flat list → wrap as "Unknown"
                if isinstance(rater_dict, list):
                    rater_dict = {"Unknown": rater_dict}

                if teacher not in old_results:
                    old_results[teacher] = {}

                for rater, docs in rater_dict.items():
                    if rater not in old_results[teacher]:
                        old_results[teacher][rater] = []

                    existing_files = {fname for fname, *_ in old_results[teacher][rater]}
                    for fname, result in docs:
                        if fname not in existing_files:
                            old_results[teacher][rater].append((fname, result))

            # merge last processed times
            old_times.update(self.last_processed_times)

            data = {
                "results": old_results,
                "last_processed_times": old_times
            }
            with open(path, "wb") as f:
                pickle.dump(data, f)

            self.processed_results = old_results
            self.last_processed_times = old_times

            print(f"✅ Results saved to {os.path.abspath(path)}")

        except Exception as e:
            print(f"❌ Error saving results: {e}")

    def _auto_export_summary(self, teacher: str):
        if not teacher:
            return

        # Ensure the controller sees freshest results
        if hasattr(self, "summary") and self.summary:
            self.summary.processed_results = self.processed_results
            try:
                self.summary.current_teacher = teacher
            except Exception:
                pass

        # Period
        ay, sem = self._infer_ay_and_sem_from_today()
        sem_label = {"1st": "1st Sem", "2nd": "2nd Sem"}.get(sem, "Summer")

        # 🔎 Resolve teacher's department (fallback if missing)
        dept = self._get_teacher_department(teacher) or "UnknownDept"

        # ✅ TARGET: Summaries/<Dept>/<AY>/<Sem>/<Teacher>/<Teacher>, <AY>, <Sem>.xlsx
        base_root = os.path.join(os.path.expanduser("~"), "Documents", "MyWork", "Summaries")

        # light sanitizing for folder/file safety
        def _safe(s: str) -> str:
            return (s or "").replace("/", "-").replace("\\", "-").strip()

        out_dir = os.path.join(base_root, _safe(dept), _safe(ay), _safe(sem), _safe(teacher))
        os.makedirs(out_dir, exist_ok=True)

        # Use the same filename style as manual export (no "Sem" word)
        target_name = f"{teacher}, {ay}, {sem}.xlsx"
        target_path = os.path.join(out_dir, _safe(target_name))

        # Prefer controller-provided writers (write directly to target_path)
        try:
            if hasattr(self.summary, "save_summary_excel"):
                self.summary.save_summary_excel(
                    target_path, teacher, include_raters=("Student", "Peer", "Self", "Dean")
                )
                protect_workbook_file(target_path)
                return
            if hasattr(self.summary, "export_summary_excel"):
                self.summary.export_summary_excel(
                    target_path, teacher, include_raters=("Student", "Peer", "Self", "Dean")
                )
                protect_workbook_file(target_path)
                return
        except Exception as ex:
            print(f"⚠️ summary helper direct-write failed: {ex}")

        # Some helpers only return a produced file path; normalize it to our target
        produced_path = None
        try:
            if hasattr(self.summary, "export_full_summary"):
                produced_path = self.summary.export_full_summary("template.xlsx")
        except Exception as ex:
            print(f"⚠️ export_full_summary failed: {ex}")
            produced_path = None

        if produced_path and os.path.exists(produced_path):
            try:
                os.makedirs(os.path.dirname(target_path), exist_ok=True)
                try:
                    os.replace(produced_path, target_path)
                except Exception:
                    shutil.copyfile(produced_path, target_path)
                    try:
                        os.remove(produced_path)
                    except Exception:
                        pass
                protect_workbook_file(target_path)
                return
            except Exception as ex:
                print(f"⚠️ Could not normalize exported excel: {ex}")

        # Fallback: minimal workbook so we still produce the correct file in the correct place
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.cell(row=1, column=1, value=f"Teacher: {teacher}")
        ws.cell(row=2, column=1, value=f"AY/Sem: {ay} / {sem_label}")

        teacher_bucket = self.processed_results.get(teacher, {}) or {}
        row = 4
        for rater in ("Student", "Peer", "Self", "Dean"):
            docs = teacher_bucket.get(rater, [])
            if not docs:
                continue
            ws.cell(row=row, column=1, value=rater); row += 1
            for fname, result_dict in docs:
                ws.cell(row=row, column=1, value=fname); row += 1
                for section, rows_map in result_dict.items():
                    ws.cell(row=row, column=1, value=section); row += 1
                    for idx, score in sorted(rows_map.items()):
                        ws.cell(row=row, column=1, value=idx)
                        ws.cell(row=row, column=2, value=score)
                        row += 1
                row += 1
            row += 1

        wb.save(target_path)
        protect_workbook_file(target_path)

    def save_csv(self):
        if not self.processed_results:
            messagebox.showwarning("Warning", "Nothing to save.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("Excel", "*.xlsx")]
        )
        if not path:
            return

        try:
            rows = []
            for teacher, rater_dict in self.processed_results.items():
                # normalize old flat format -> wrap as Unknown
                if isinstance(rater_dict, list):
                    rater_dict = {"Unknown": rater_dict}

                for rater, docs in rater_dict.items():
                    for file, result in docs:
                        row = {"Teacher": teacher, "Rater": rater, "File": file}
                        for sec, sec_data in result.items():
                            for rownum, score in sec_data.items():
                                row[f"{sec} Row {rownum}"] = score
                        rows.append(row)

            df = pd.DataFrame(rows)
            if path.endswith(".csv"):
                df.to_csv(path, index=False)
            else:
                df.to_excel(path, index=False)

            messagebox.showinfo("Saved", f"Results saved to {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    def _next_scan_index(self, folder: str) -> int:
        """Return next integer N so the next file can be named ..._{N:03d}.bmp."""
        import re
        max_n = 0
        if not os.path.exists(folder):
            return 1
        for entry in os.scandir(folder):
            if not entry.name.lower().endswith(".bmp"):
                continue
            # grab the last number before .bmp (works with page_12.bmp or Teacher_012.bmp)
            m = re.search(r'(\d+)(?=\.bmp$)', entry.name, flags=re.IGNORECASE)
            if m:
                max_n = max(max_n, int(m.group(1)))
        return max_n + 1

     
    def load_results(self, path=None):
        # 🔹 default path = same folder as database
        if path is None:
            db_path = db.get_default_db_path()
            base_dir = os.path.dirname(db_path)
            path = os.path.join(base_dir, "results.pkl")

        if not os.path.exists(path):
            print("⚠️ No saved results found.")
            self.processed_results = {}
            self.last_processed_times = {}
            return

        try:
            with open(path, "rb") as f:
                data = pickle.load(f)

            self.processed_results = data.get("results", {})
            self.last_processed_times = data.get("last_processed_times", {})

            # 🔑 Normalize: wrap flat lists into {"Unknown": [...]}
            for teacher, val in list(self.processed_results.items()):
                if isinstance(val, list):
                    self.processed_results[teacher] = {"Unknown": val}

            print(f"✅ Results loaded from {os.path.abspath(path)}")
        except Exception as e:
            print(f"❌ Error loading results: {e}")
            self.processed_results = {}
            self.last_processed_times = {}

   
    #---------------- PREVIEW HANDLERS ---------------- #
    def update_preview(self, teacher_results):
        if not teacher_results:
            self.document_listbox.configure(values=["No documents loaded"])
            self.document_listbox.set("No documents loaded")
            self.img_label.configure(image=None, text="No image loaded")
            return

        values = [fname for fname, *_ in teacher_results]
        self.document_listbox.configure(values=values)
        self.document_listbox.set(values[0])

        # ensure command is bound after repopulating
        self.document_listbox.configure(
            command=lambda choice: self.display_image(
                choice, self.teacher_var.get(), base_dir=self.current_scan_dir
            )
        )

        # load the first image immediately
        self.display_image(values[0], self.teacher_var.get(), base_dir=self.current_scan_dir)

    def display_image(self, filename, teacher, base_dir=None):
        try:
            folder = base_dir or self.current_scan_dir or self._build_scan_dir(teacher)
            full_path = os.path.join(folder, filename)

            if full_path in self.annotated_cache:
                rgb_img = cv2.cvtColor(self.annotated_cache[full_path], cv2.COLOR_BGR2RGB)
                pil_img = Image.fromarray(rgb_img).resize((400, 500), Image.Resampling.LANCZOS)
            else:
                pil_img = Image.open(full_path).resize((400, 500), Image.Resampling.LANCZOS)

            img_tk = CTkImage(light_image=pil_img, dark_image=pil_img, size=(400, 500))
            self.img_label.configure(image=img_tk, text="")
            self.img_label.image = img_tk  # prevent GC
        except Exception as e:
            self.img_label.configure(image=None, text="Error loading image")
            messagebox.showerror("Image Error", str(e))

    # ---------------- OTHERS ---------------- #
  
    def set_controls_state(self, state="normal"):
        """Enable or disable all buttons/menus in the scan page."""
        widgets = [
            getattr(self, "btn_scan", None),
            getattr(self, "btn_check_existing", None),
            getattr(self, "btn_clear_docs", None),
            getattr(self, "teacher_dropdown", None),
            getattr(self, "subject_dropdown", None),
            getattr(self, "rater_dropdown", None),
            getattr(self, "document_listbox", None),
        ]
        for w in widgets:
            try:
                if w is not None:
                    w.configure(state=state)
            except Exception:
                pass

    def _get_expected_total_for_teacher(self, teacher_full_name: str) -> int:
        """Sum expected students across this term's teaching_loads for the teacher."""
        try:
            ay, sem = self._infer_ay_and_sem_from_today()
            with db.connect() as conn:
                row = conn.execute("""
                    SELECT COALESCE(SUM(tl.expected_students), 0)
                    FROM teaching_loads tl
                    JOIN faculty f ON f.id = tl.teacher_id
                    JOIN curriculum_subjects cs ON cs.id = tl.curriculum_subject_id
                    WHERE f.full_name = ?
                    AND cs.academic_year = ?
                    AND cs.semester = ?
                """, (teacher_full_name, ay, sem)).fetchone()
            return int(row[0] or 0)
        except Exception:
            return 0

    def _get_completed_student_scans_current_period(self, teacher_full_name: str) -> int:
        """
        Count Student scans for this teacher that are saved and belong to the CURRENT AY/Sem folder.
        We ensure the file exists under the current scan dir to scope to the period.
        """
        # where results.pkl is stored
        db_path = db.get_default_db_path()
        base_dir = os.path.dirname(db_path)
        pkl_path = os.path.join(base_dir, "results.pkl")

        if not os.path.exists(pkl_path):
            return 0

        try:
            with open(pkl_path, "rb") as f:
                data = pickle.load(f)
            results = data.get("results", {})
        except Exception:
            return 0

        teacher_bucket = results.get(teacher_full_name, {})
        student_list = teacher_bucket.get("Student", [])
        if isinstance(teacher_bucket, list):
            # normalize legacy structure (shouldn’t happen here, but safe)
            student_list = []

        # Only count files that live in THIS period's folder to avoid cross-term inflation
        current_dir = self._build_scan_dir(teacher_full_name)
        count = 0
        for fname, _payload in student_list:
            if os.path.exists(os.path.join(current_dir, fname)):
                count += 1
        return count

    def _refresh_teacher_progress(self):
        teacher = (self.teacher_var.get() or "").strip()
        if not teacher or teacher in ("Loading...", "No teachers found"):
            self.progress_bar.set(0)
            self.scan_info_label.configure(text="No teacher selected")
            return

        subj_label = (self.subject_var.get() or "").strip()
        subj_code  = self.subject_code_by_label.get(subj_label, "")
        if not subj_code:
            self.progress_bar.set(0)
            self.scan_info_label.configure(text="Pick a subject for this teacher")
            return

        total = self._get_expected_total_for_teacher_subject(teacher, subj_code)
        completed = self._get_completed_student_scans_current_period_subject(teacher, subj_code)
        remaining = max(total - completed, 0)

        self.progress_bar.set((completed / total) if total > 0 else 0.0)
        self.scan_info_label.configure(text=f"{completed} / {total} completed • {remaining} remaining")


    def _on_teacher_change(self):
        self._load_subjects_for_teacher(self.teacher_var.get())
        self._refresh_teacher_progress()

    def _load_subjects_for_teacher(self, teacher_full_name: str):
        """Fill the Subject dropdown with this teacher's subjects this term."""
        try:
            ay, sem = self._infer_ay_and_sem_from_today()
            with db.connect() as conn:
                rows = conn.execute("""
                    SELECT s.id, s.code, s.title
                    FROM teaching_assignments ta
                    JOIN subjects s ON s.id = ta.subject_id
                    JOIN faculty  f ON f.id = ta.teacher_id
                    WHERE f.full_name = ? AND ta.academic_year = ? AND ta.semester = ?
                    GROUP BY s.id, s.code, s.title
                    ORDER BY s.code
                """, (teacher_full_name, ay, sem)).fetchall()

            if not rows:
                self.subject_dropdown.configure(values=["— No subjects this term —"], state="disabled")
                self.subject_var.set("— No subjects this term —")
                self.subject_label_to_id.clear()
                self.subject_code_by_label.clear()
                return

            labels = []
            self.subject_label_to_id.clear()
            self.subject_code_by_label.clear()
            for sid, code, title in rows:
                label = f"{code} — {title}"
                labels.append(label)
                self.subject_label_to_id[label] = sid
                self.subject_code_by_label[label] = code

            self.subject_dropdown.configure(values=labels, state="normal")
            self.subject_var.set(labels[0])

        except Exception as e:
            messagebox.showerror("DB Error", str(e))
            self.subject_dropdown.configure(values=["— Error —"], state="disabled")
            self.subject_var.set("— Error —")

    def _meta_path(self, teacher_root: str) -> str:
        return os.path.join(teacher_root, "scan_meta.json")

    def _load_subject_meta(self, teacher_root: str) -> dict:
        try:
            p = self._meta_path(teacher_root)
            if os.path.exists(p):
                with open(p, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        return {}

    def _save_subject_meta(self, teacher_root: str, filename: str, subject_code: str):
        try:
            meta = self._load_subject_meta(teacher_root)
            meta[filename] = subject_code or ""
            with open(self._meta_path(teacher_root), "w", encoding="utf-8") as f:
                json.dump(meta, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
 
    def _get_expected_total_for_teacher_subject(self, teacher_full_name: str, subject_code: str) -> int:
        if not subject_code:
            return 0
        try:
            ay, sem = self._infer_ay_and_sem_from_today()
            with db.connect() as conn:
                row = conn.execute("""
                    SELECT COALESCE(SUM(ta.expected_students), 0)
                    FROM teaching_assignments ta
                    JOIN subjects s ON s.id = ta.subject_id
                    JOIN faculty  f ON f.id = ta.teacher_id
                    WHERE f.full_name = ? AND s.code = ? AND ta.academic_year = ? AND ta.semester = ?
                """, (teacher_full_name, subject_code, ay, sem)).fetchone()
            return int(row[0] or 0)
        except Exception:
            return 0

    def _get_completed_student_scans_current_period_subject(self, teacher_full_name: str, subject_code: str) -> int:
        """Count Student scans for this teacher in THIS AY/Sem that were tagged with subject_code."""
        if not subject_code:
            return 0

        db_path = db.get_default_db_path()
        base_dir = os.path.dirname(db_path)
        pkl_path = os.path.join(base_dir, "results.pkl")
        if not os.path.exists(pkl_path):
            return 0

        try:
            with open(pkl_path, "rb") as f:
                data = pickle.load(f)
            results = data.get("results", {})
        except Exception:
            return 0

        teacher_bucket = results.get(teacher_full_name, {})
        student_list = teacher_bucket.get("Student", [])
        if isinstance(teacher_bucket, list):
            student_list = []

        current_dir = self._build_scan_dir(teacher_full_name)
        meta = self._load_subject_meta(current_dir)

        count = 0
        for fname, _payload in student_list:
            if os.path.exists(os.path.join(current_dir, fname)) and meta.get(fname, "") == subject_code:
                count += 1
        return count
    

       
                
