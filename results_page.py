from customtkinter import *
from CTkTable import *
from tkinter import filedialog, messagebox
import os
import sys
import pickle
from tkinter import ttk
from datetime import date
import re
import shutil, time
from pathlib import Path
from openpyxl import load_workbook

import db
import utils
from summary_helpers import SummaryFormController, protect_workbook_file   # <-- NEW


class ResultsPage:
    def __init__(self, master, processed_results):
        self.master = master
        self.processed_results = processed_results
        self.current_teacher = None
        self.tab_buttons = {}

        # Load results from pickle
        self.load_results()
        # Prepare shared Summary controller (module)
        self.summary = SummaryFormController(self.processed_results, db_module=db)

        self._build_ui()
    def _show_archive_dialog(self, zip_path: Path, department: str, academic_year: str, semester: str):
        """
        ATS-themed dialog to notify that the archive was successfully created.
        """
        TOPBAR = "#BF3131"
        SIDEBAR_BTN = "#AC5353"
        HOVER = "#BF3131"
        PANEL_BG = "#F5F5F5"
        LIGHT_TEXT = "#FFEFEF"
        WHITE = "#FFFFFF"

        parent = self.master

        dialog = CTkToplevel(parent)
        dialog.title("Archive Created")
        dialog.resizable(False, False)

        dialog.transient(parent)
        dialog.grab_set()

        main_frame = CTkFrame(dialog, fg_color=PANEL_BG, corner_radius=12)
        main_frame.pack(fill="both", expand=True, padx=16, pady=16)

        # Header bar
        header = CTkFrame(main_frame, fg_color=TOPBAR, corner_radius=10)
        header.pack(fill="x", padx=8, pady=(8, 4))

        CTkLabel(
            header,
            text="Archive successfully created",
            font=("Poppins", 16, "bold"),
            text_color=WHITE
        ).pack(side="left", padx=12, pady=8)

        CTkLabel(
            header,
            text="ARCHIVE",
            font=("Poppins", 11, "bold"),
            text_color=TOPBAR,
            fg_color=LIGHT_TEXT,
            corner_radius=999,
            padx=10,
            pady=4,
        ).pack(side="right", padx=12, pady=8)

        # Body
        body = CTkFrame(main_frame, fg_color=WHITE, corner_radius=10)
        # 🔧 don't let body steal *all* vertical space
        body.pack(fill="both", expand=False, padx=8, pady=(4, 8))

        CTkLabel(
            body,
            text=(
                "Your semester archive has been generated with the following details:\n\n"
                f"• Department: {department}\n"
                f"• Academic Year: {academic_year}\n"
                f"• Semester: {semester}\n\n"
                "Archive file:"
            ),
            font=("Poppins", 12),
            text_color="#333333",
            justify="left",
        ).pack(anchor="w", padx=12, pady=(12, 4))

        # Path label (slightly smaller / dimmer)
        CTkLabel(
            body,
            text=str(zip_path),
            font=("Poppins", 11),
            text_color="#6B7280",
            justify="left"
        ).pack(anchor="w", padx=12, pady=(0, 12))

        # Footer buttons
        footer = CTkFrame(main_frame, fg_color=PANEL_BG)
        footer.pack(fill="x", padx=8, pady=(0, 8))

        def open_folder():
            try:
                folder = zip_path.parent
                if os.name == "nt":
                    os.startfile(folder)
                elif sys.platform == "darwin":
                    import subprocess
                    subprocess.run(["open", folder])
                else:
                    import subprocess
                    subprocess.run(["xdg-open", folder])
            except Exception:
                # swallow errors silently, dialog stays open
                pass

        def close_dialog():
            dialog.destroy()

        CTkButton(
            footer,
            text="Close",
            font=("Poppins", 12, "bold"),
            fg_color="#E5E7EB",
            hover_color="#D1D5DB",
            text_color="#111827",
            corner_radius=8,
            width=120,
            command=close_dialog
        ).pack(side="right", padx=8, pady=4)

        CTkButton(
            footer,
            text="Open Folder",
            font=("Poppins", 12, "bold"),
            fg_color=SIDEBAR_BTN,
            hover_color=HOVER,
            text_color=WHITE,
            corner_radius=8,
            width=140,
            command=open_folder
        ).pack(side="right", padx=8, pady=4)

        # Center dialog over parent – AFTER layout so we know its size
        dialog.update_idletasks()
        w = dialog.winfo_width()
        h = dialog.winfo_height()

        # Optional: enforce a minimum width/height just in case
        if w < 520:
            w = 520
        if h < 260:
            h = 260
        dialog.minsize(w, h)

        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (w // 2)
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (h // 2)
        dialog.geometry(f"{w}x{h}+{x}+{y}")

        dialog.wait_window(dialog)


    # ---------------- UI ---------------- #
    def _build_ui(self):
        for w in self.master.winfo_children():
            w.destroy()

        self.container = CTkFrame(self.master, fg_color="transparent")
        self.container.pack(fill="both", expand=True, padx=20, pady=20)

        # --- Page Title Bar (same style as Data Management Panel) ---
        title_bar = CTkFrame(self.container, fg_color="#BF3131", height=70, corner_radius=10)
        title_bar.pack(fill="x", padx=10, pady=(0, 12))

        CTkLabel(
            title_bar,
            text="Evaluation Results",          # <- you can change this text if you want
            font=("Poppins", 18, "bold"),
            text_color="#FFFFFF"
        ).pack(side="left", padx=20, pady=12)

        # Tabs: teachers
        self.tab_frame = CTkFrame(self.container, fg_color="transparent")
        self.tab_frame.pack(fill="x", pady=(0, 10))

        self.content_frame = CTkFrame(self.container, fg_color="#FFFFFF", corner_radius=10)
        self.content_frame.pack(fill="both", expand=True)

        self.controls_frame = CTkFrame(self.container, fg_color="transparent")
        self.controls_frame.pack(fill="x", pady=(10, 0))

        self._build_tabs()

    # ---------------- Tabs ---------------- #
    def _teacher_department_id(self, teacher_name: str) -> int | None:
        if not teacher_name:
            return None
        try:
            with db.connect() as conn:
                row = conn.execute(
                    "SELECT department_id FROM faculty WHERE full_name = ?",
                    (teacher_name,),
                ).fetchone()
                return row[0] if row else None
        except Exception:
            return None

    def _build_tabs(self):
        self.tab_buttons = {}
        first_teacher = None
        user_dept_id = utils.get_current_user().get("department_id")

        # Filter teachers by current user's department if specified
        all_teachers = list(self.processed_results.keys())
        if user_dept_id is not None:
            teachers = [
                t for t in all_teachers
                if self._teacher_department_id(t) == user_dept_id
            ]
            # if no matches, fall back to all to avoid empty UI
            if not teachers:
                teachers = all_teachers
        else:
            teachers = all_teachers

        for teacher in teachers:
            if not first_teacher:
                first_teacher = teacher
            btn = CTkButton(
                self.tab_frame,
                text=teacher,
                command=lambda t=teacher: self.show_teacher(t),
                fg_color="#AC5353",          # default inactive color
                hover_color="#8B1D18",       # Dark Red hover for inactive
                text_color="#333333",
                width=160, height=35, corner_radius=8
            )
            btn.pack(side="left", padx=5)
            self.tab_buttons[teacher] = btn

        if first_teacher:
            self.show_teacher(first_teacher)

    def show_teacher(self, teacher):
        self.current_teacher = teacher
        # Reset tab highlight with proper palette alignment
        for t, btn in self.tab_buttons.items():
            if t == teacher:
                btn.configure(
                    fg_color="#691612",    # Dark Crimson active
                    text_color="#FFFFFF",  # White text active
                    hover_color="#BF3131"  # Crimson hover active
                )
            else:
                btn.configure(
                    fg_color="#AC5353",    # Muted Red inactive
                    text_color="#333333",  # Dark text inactive
                    hover_color="#8B1D18"  # Dark Red hover inactive
                )

        # Build rater dropdown + table
        for w in self.content_frame.winfo_children():
            w.destroy()

        header_frame = CTkFrame(self.content_frame, fg_color="transparent")
        header_frame.pack(fill="x", padx=20, pady=10)

        CTkLabel(
            header_frame,
            text=f"Results for {teacher}",
            font=("Poppins", 20, "bold"),
            text_color="#691612"
        ).pack(side="left")

        rater_options = list(self.processed_results[teacher].keys())
        self.rater_var = StringVar(value=rater_options[0])
        CTkOptionMenu(
            header_frame,
            variable=self.rater_var,
            values=rater_options,
            command=lambda r: self.build_table(r),
            fg_color="#691612",
            button_color="#AC5353",
            text_color="#FFFFFF",
            width=180
        ).pack(side="right")

        self.table_container = CTkFrame(self.content_frame, fg_color="transparent")
        self.table_container.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        # Controls container (bottom bar)
        for w in self.controls_frame.winfo_children():
            w.destroy()
        self._build_controls(self.controls_frame)

        # Default rater = first available
        if rater_options:
            self.rater_var.set(rater_options[0])
            self.build_table(rater_options[0])

    # ---------------- Table ---------------- #
    def build_table(self, rater_type):
        self.current_rater = rater_type
        for w in self.table_container.winfo_children():
            w.destroy()

        data = self.processed_results[self.current_teacher].get(rater_type, [])
        if not data:
            CTkLabel(
                self.table_container,
                text="No results available",
                font=("Poppins", 14),
                text_color="#1F2937"
            ).pack(pady=20)
            return

        # ---------- Section Titles ----------
        section_titles = {
            "Section 1": "I. Commitment",
            "Section 2": "II. Knowledge of Subject",
            "Section 3": "III. Teaching for Independent Learning",
            "Section 4": "IV. Management of Learning"
        }

        # ---------- Headers ----------
        headers = ["Question"]
        for idx, (fname, result, *_) in enumerate(data):
            headers.append(f"{fname}")

        # ---------- Collect rows ----------
        table_data = []
        sections = sorted({sec for _, result, *_ in data for sec in result.keys()})
        for sec in sections:
            title = section_titles.get(sec, sec)
            table_data.append([title] + [""] * len(data))
            max_rows = max(len(result.get(sec, {})) for _, result, *_ in data)
            for rownum in range(1, max_rows + 1):
                row = [f"{rownum}"]
                for _, result, *_ in data:
                    score = result.get(sec, {}).get(rownum, "")
                    row.append(score)
                table_data.append(row)

        # ---------- Treeview ----------
        frame = CTkFrame(self.table_container, fg_color="transparent")
        frame.pack(fill="both", expand=True)

        x_scroll = CTkScrollbar(frame, orientation="horizontal")
        x_scroll.pack(side="bottom", fill="x")
        y_scroll = CTkScrollbar(frame, orientation="vertical")
        y_scroll.pack(side="right", fill="y")

        self.tree = ttk.Treeview(
            frame,
            columns=headers,
            show="headings",
            xscrollcommand=x_scroll.set,
            yscrollcommand=y_scroll.set,
            height=20
        )
        self.tree.pack(fill="both", expand=True)

        x_scroll.configure(command=self.tree.xview)
        y_scroll.configure(command=self.tree.yview)

        # ---------- Styling ----------
        style = ttk.Style()
        style.configure(
            "Treeview",
            font=("Poppins", 11),
            rowheight=30,
            background="#F3F4F6",
            foreground="#1F2937"
        )
        style.configure(
            "Treeview.Heading",
            font=("Poppins", 12, "bold"),
            foreground="#FFFFFF",
            background="#BF3131"
        )
        style.map(
            "Treeview",
            background=[("selected", "#AC5353")],
            foreground=[("selected", "#FFFFFF")]
        )
        style.map(
            "Treeview.Heading",
            background=[("hover", "#AC5353")],
            foreground=[("hover", "#FFFFFF")]
        )

        self.tree.tag_configure("oddrow", background="#F3F4F6", foreground="#1F2937")
        self.tree.tag_configure("evenrow", background="#F3F4F6", foreground="#1F2937")
        self.tree.tag_configure(
            "section",
            background="#BF3131",
            foreground="#FFFFFF",
            font=("Poppins", 12, "bold")
        )

        # ---------- Insert Data ----------
        for h in headers:
            self.tree.heading(h, text=h)
            self.tree.column(h, width=120, anchor="center")

        section_row_index = 0
        for row in table_data:
            if row[0].startswith(("I.", "II.", "III.", "IV.")):
                self.tree.insert("", "end", values=row, tags=("section",))
                section_row_index = 0
            else:
                tag = "evenrow" if section_row_index % 2 == 0 else "oddrow"
                self.tree.insert("", "end", values=row, tags=(tag,))
                section_row_index += 1

    # ---------------- Controls ---------------- #

    def _build_controls(self, parent):
        control_frame = CTkFrame(parent, fg_color="transparent")
        control_frame.pack(fill="x", padx=10, pady=10)


        # ----- Common button style -----
        def make_button(master, text, command, fg="#AC5353", hover="#BF3131", width=170):
            return CTkButton(
                master,
                text=text,
                command=command,
                fg_color=fg,
                hover_color=hover,
                text_color="#FFFFFF",
                font=("Poppins", 14, "bold"),
                height=38,
                corner_radius=8,
                width=width,
            )

        # Primary actions (left group)
        btn_group_left = CTkFrame(control_frame, fg_color="transparent")
        btn_group_left.pack(side="left")

        export_btn = make_button(
            btn_group_left,
            text="Export Summary",
            command=lambda: self._export_summary_via_module(),
        )
        export_btn.pack(side="left", padx=5)

        view_summaries_btn = make_button(
            btn_group_left,
            text="View Summaries",
            command=self._open_summaries_folder,
            width=160,
        )
        view_summaries_btn.pack(side="left", padx=5)

        create_summary_btn = make_button(
            btn_group_left,
            text="Create Summary",
            command=self._create_final_summary,
            width=160,
        )
        create_summary_btn.pack(side="left", padx=5)

        # Spacer to push Archive to the right
        CTkFrame(control_frame, fg_color="transparent").pack(side="left", expand=True)

        # Archive (stronger / darker accent, right aligned)
        archive_btn = make_button(
            control_frame,
            text="Archive this Semester",
            command=self.archive_current_semester,
            fg="#691612",
            hover="#8B1D18",
            width=210,
        )
        archive_btn.pack(side="right", padx=5)


    def _open_summary_for_current(self):
        if not self.current_teacher:
            messagebox.showwarning("No Teacher", "Please select a teacher first.")
            return
        self.summary.show(self.master, self.current_teacher)

    def _export_summary_via_module(self):
        if not self.current_teacher:
            messagebox.showwarning("No Teacher", "Please select a teacher first.")
            return
        try:
            # ensure controller has fresh context
            self.summary.processed_results = self.processed_results
            self.summary.current_teacher = self.current_teacher

            save_path = self.summary.export_full_summary("template.xlsx")
            if save_path:
                protect_workbook_file(save_path)
                messagebox.showinfo("Saved", f"✅ Summary updated:\n{save_path}")

                # Activity log (kept from your original export)
                user = utils.get_current_user()
                filename = os.path.basename(save_path)
                # Try to parse AY/Sem from filename for the log (best-effort)
                details = {"path": save_path}
                try:
                    parts = os.path.splitext(filename)[0].split(", ")
                    if len(parts) >= 3:
                        details["academic_year"] = parts[-2]
                        details["semester"] = parts[-1]
                except Exception:
                    pass

                db.log_activity(
                    action="export_summary",
                    actor_name=user.get("name"),
                    actor_role=user.get("role"),
                    department_id=user.get("department_id"),
                    teacher_name=self.current_teacher,
                    file_name=filename,
                    details=details
                )
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    def _get_department_name(self):
        user = utils.get_current_user()
        dept_id = user.get("department_id") if isinstance(user, dict) else None
        if dept_id is None:
            return None
        try:
            with db.connect() as conn:
                row = conn.execute("SELECT name FROM departments WHERE id=?", (dept_id,)).fetchone()
                if row:
                    return row[0]
        except Exception:
            return None
        return None

    def _get_summary_folder(self):
        dept_name = self._get_department_name()
        if not dept_name:
            return None
        ay, sem = self._infer_ay_and_sem_from_today()
        base_root = Path(os.path.join(os.path.expanduser("~"), "Documents", "MyWork", "Summaries"))
        return base_root / dept_name / ay / sem

    def _open_summaries_folder(self):
        """Open the Summaries folder for current dept/AY/Sem."""
        target = self._get_summary_folder()
        if target is None:
            messagebox.showwarning("Missing Department", "Cannot determine your department.")
            return
        if not target.exists():
            messagebox.showwarning("Not Found", f"Folder not found:\n{target}")
            return

        try:
            if os.name == "nt":
                os.startfile(target)
            elif sys.platform == "darwin":
                import subprocess
                subprocess.run(["open", str(target)])
            else:
                import subprocess
                subprocess.run(["xdg-open", str(target)])
        except Exception as ex:
            messagebox.showerror("Open Folder Failed", str(ex))

    def _create_final_summary(self):
        """Generate the final summary directly for the current dept/AY/Sem (no dialog)."""
        sem_folder = self._get_summary_folder()
        if sem_folder is None:
            messagebox.showwarning("Missing Department", "Cannot determine your department.")
            return
        try:
            sem_folder.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            messagebox.showerror("Folder Error", f"Could not prepare folder:\n{sem_folder}\n\n{exc}")
            return

        # Template lives beside this file (works even if cwd differs)
        template_path = Path(__file__).resolve().parent / "final_summary.xlsx"
        self._create_final_summary_for_semester(str(sem_folder), str(template_path))


    def _rating_to_adjective(self,score):
        """Convert numeric rating to adjective."""
        if score is None:
            return ""
        try:
            s = float(score)
        except Exception:
            return ""
        if s >= 9.3:
            return "Outstanding"
        if s >= 7.5:
            return "Very Satisfactory"
        if s >= 5.0:
            return "Satisfactory"
        if s >= 3.0:
            return "Fair"
        if s >= 2.0:
            return "Unsatisfactory"
        return "Unsatisfactory"

    def _create_final_summary_for_semester(self, sem_folder: str, template_path: str):
        """
        sem_folder: the folder 'Summaries > Dept > AY > Sem'
        template_path: path to final_summary.xlsx template
        """
        sem_path = Path(sem_folder)
        template = Path(template_path)

        if not sem_path.exists():
            messagebox.showerror("Error", f"Semester folder not found:\n{sem_path}")
            return
        if not template.exists():
            messagebox.showerror("Error", f"Template not found:\n{template}")
            return

        # 1) Collect all teacher summaries (one per subfolder)
        teacher_rows = []  # (name, rating, adjective)

        for teacher_dir in sorted(sem_path.iterdir()):
            if not teacher_dir.is_dir():
                continue

            # find the teacher's summary file inside this folder
            candidates = [
                p for p in teacher_dir.glob("*.xlsx")
                if p.name.lower() not in ("final_summary.xlsx", "template.xlsx")
                and not p.name.startswith("~$")
            ]
            if not candidates:
                continue

            # If multiple, pick the newest
            candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            src = candidates[0]

            try:
                # open teacher summary and read Faculty!C9 and Faculty!S37
                wb_teacher = load_workbook(src, data_only=True)
                ws_fac = wb_teacher["Faculty"]  # adjust if your sheet name differs

                name = ws_fac["C9"].value
                rating = ws_fac["S37"].value
                adjective = self._rating_to_adjective(rating)

                if not name:
                    # fallback to folder name if C9 is empty
                    name = teacher_dir.name

                teacher_rows.append((name, rating, adjective))
            except Exception as e:
                messagebox.showwarning(
                    "Warning",
                    f"Could not read summary for {teacher_dir.name}:\n{e}"
                )

        if not teacher_rows:
            messagebox.showinfo("No Data", "No teacher summaries found in this semester folder.")
            return

        # 2) Create a copy of final_summary.xlsx in the sem folder
        dest_path = sem_path / "final_summary.xlsx"
        shutil.copyfile(template, dest_path)

        # 3) Open that copy and write the rows
        wb_final = load_workbook(dest_path, data_only=False)
        # case-insensitive sheet lookup just in case
        sheet_name = None
        for s in wb_final.sheetnames:
            if s.lower() == "summary":
                sheet_name = s
                break
        if sheet_name is None:
            messagebox.showerror("Error", "No 'Summary' sheet found in final_summary.xlsx template.")
            return

        ws_sum = wb_final[sheet_name]

        start_row = 12  # rows start at 12
        for idx, (name, rating, adjective) in enumerate(teacher_rows):
            row = start_row + idx

            # B column: name
            c_name = ws_sum.cell(row=row, column=2)   # B
            c_name.value = str(name) if name is not None else ""

            # I column: rating
            c_rating = ws_sum.cell(row=row, column=9)  # I
            try:
                c_rating.value = float(rating) if rating is not None else None
            except Exception:
                c_rating.value = rating

            # O column: adjective
            c_adj = ws_sum.cell(row=row, column=15)   # O
            c_adj.value = str(adjective) if adjective is not None else ""

        # 4) (Optional but recommended) clear any leftover formulas in B/I/O below our data
        max_rows_to_clean = 200  # adjust if you can have more than 188 teachers
        end_clean_row = start_row + max_rows_to_clean
        last_used_row = start_row + len(teacher_rows)

        for row in range(last_used_row, end_clean_row + 1):
            for col in (2, 9, 15):  # B, I, O
                cell = ws_sum.cell(row=row, column=col)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = None  # remove TI!DC6, etc.

        wb_final.save(dest_path)
        messagebox.showinfo("Summary Created", f"Final summary saved:\n{dest_path}")

    
    # ---------------- Data Persistence ---------------- #
    def load_results(self, path=None):
        """Load processed_results dict from a pickle file (teacher → rater → docs)."""
        # default path = same folder as database
        if path is None:
            db_path = db.get_default_db_path()
            base_dir = os.path.dirname(db_path)
            path = os.path.join(base_dir, "results.pkl")

        if not os.path.exists(path):
            print("⚠️ No saved results found.")
            self.processed_results = {}
            return {}

        try:
            with open(path, "rb") as f:
                data = pickle.load(f)

            # Ensure we extract results correctly
            if isinstance(data, dict) and "results" in data:
                self.processed_results = data["results"]
            else:
                # legacy format fallback: wrap flat structure into "Unknown" rater
                self.processed_results = {}
                for teacher, docs in data.items():
                    self.processed_results[teacher] = {"Unknown": docs}

            return self.processed_results

        except Exception as e:
            print(f"❌ Error loading results: {e}")
            self.processed_results = {}
            return {}

    # ---------------- Static Helpers ---------------- #
    @staticmethod
    def _infer_ay_and_sem_from_today():
        """Infer AY and Sem from today's date (PH academic calendar)."""
        today = date.today()
        y, m = today.year, today.month
        if 8 <= m <= 12:
            return f"{y}-{y+1}", "1st Sem"
        elif 1 <= m <= 5:
            return f"{y-1}-{y}", "2nd Sem"
        else:  # Jun–Jul
            return f"{y-1}-{y}", "Midyear"

    # ---------------- Archive Semester ---------------- #
    def archive_current_semester(self):
        """
        Zip a selected semester folder (department > academic year > semester > profs)
        and store to: Archived > department > academic year > <semester>_<timestamp>.zip
        """
        # Prefill the dialog to the Scan/<Dept> root to save clicks
        try:
            base_dir = Path(os.path.join(os.path.expanduser("~"), "Documents", "MyWork"))
            scan_root = base_dir / "Scan"
            initial_dir = scan_root
            user = utils.get_current_user()
            dept_id = user.get("department_id") if isinstance(user, dict) else None
            if dept_id is not None:
                with db.connect() as conn:
                    row = conn.execute("SELECT name FROM departments WHERE id=?", (dept_id,)).fetchone()
                if row and row[0]:
                    candidate = scan_root / row[0]
                    if candidate.exists():
                        initial_dir = candidate
        except Exception:
            initial_dir = None

        semester_dir = filedialog.askdirectory(
            title="Select the SEMESTER folder to archive (e.g., .../Department/2025-2026/1st)",
            initialdir=str(initial_dir) if initial_dir else None,
        )
        if not semester_dir:
            return

        sem_path = Path(semester_dir)
        if not sem_path.exists() or not sem_path.is_dir():
            messagebox.showerror("Invalid folder", "Please select a valid semester folder.")
            return

        # Expect path like .../<Department>/<Academic Year>/<Semester>
        try:
            semester = sem_path.name
            academic_year = sem_path.parent.name
            department = sem_path.parent.parent.name
        except Exception:
            messagebox.showerror(
                "Path error",
                "Could not infer Department/Academic Year/Semester from the selected path.\n"
                "Expected: .../<Department>/<Academic Year>/<Semester>"
            )
            return

        # Resolve MyWork root using your DB helper
        try:
            base_dir = Path(os.path.dirname(db.get_default_db_path()))
        except Exception:
            base_dir = Path(os.path.expanduser("~")) / "Documents" / "MyWork"

        # Build archive destination: .../Archived/<Department>/<Academic Year>/
        archive_root = base_dir / "Archived" / department / academic_year
        archive_root.mkdir(parents=True, exist_ok=True)

        # ZIP name: <Semester>_<YYYYMMDD-HHMM>.zip
        stamp = time.strftime("%Y%m%d-%H%M")
        base_name = f"{semester}_{stamp}"
        zip_noext = archive_root / base_name

        try:
            shutil.make_archive(str(zip_noext), 'zip', root_dir=str(sem_path))
        except Exception as e:
            messagebox.showerror("Archive failed", f"Could not create archive:\n{e}")
            return

        # Use the custom ATS-styled dialog instead of a basic messagebox
        self._show_archive_dialog(
            zip_path=zip_noext.with_suffix(".zip"),
            department=department,
            academic_year=academic_year,
            semester=semester
        )
